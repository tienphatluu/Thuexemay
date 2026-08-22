# -*- coding: utf-8 -*-
import os,sqlite3,tkinter as tk
from tkinter import ttk,messagebox,filedialog
from datetime import datetime,timedelta
from tkcalendar import DateEntry

try:
    from PIL import Image,ImageDraw,ImageFont
    PIL_OK=True
except ImportError:PIL_OK=False
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font,PatternFill,Alignment
    from openpyxl.utils import get_column_letter
    XLSX_OK=True
except ImportError:XLSX_OK=False

BASE=os.path.dirname(os.path.abspath(__file__))
DB_FILE=os.path.join(BASE,"rental_v2.db")
PHIEU_DIR=os.path.join(BASE,"PHIEU_GIAO_XE")
QR_FILE=os.path.join(BASE,"QR_THANH_TOAN.png")
FMT="%d/%m/%Y %H:%M"
os.makedirs(PHIEU_DIR,exist_ok=True)

def now():return datetime.now()
def money(v):
    try:return f"{float(v):,.0f}".replace(",",".")+" đ"
    except:return "0 đ"
def num(v):
    try:return float(v)
    except:return 0
def money_input(v):
    s=str(v).strip().replace(" ","").replace("đ","").replace("Đ","").replace(".","").replace(",","")
    if not s:return 0
    try:n=float(s)
    except:raise ValueError("Số tiền không hợp lệ.")
    if n<0:raise ValueError("Số tiền không được âm.")
    return n
def dt(v):return datetime.strptime(v.strip(),FMT)
def hours_list():return [f"{h:02d}:{m:02d}" for h in range(24) for m in (0,15,30,45)]

class DB:
    def __init__(self):
        self.c=sqlite3.connect(DB_FILE);self.c.row_factory=sqlite3.Row
        self.create();self.migrate_money()
    def create(self):
        self.c.executescript("""
        CREATE TABLE IF NOT EXISTS vehicles(
            id INTEGER PRIMARY KEY AUTOINCREMENT,code TEXT UNIQUE NOT NULL,
            plate TEXT UNIQUE NOT NULL,model TEXT,color TEXT,
            default_price REAL DEFAULT 0,status TEXT DEFAULT 'Trống',note TEXT);
        CREATE TABLE IF NOT EXISTS customers(
            id INTEGER PRIMARY KEY AUTOINCREMENT,name TEXT NOT NULL,phone TEXT,
            id_number TEXT,address TEXT,note TEXT,created_at TEXT);
        CREATE TABLE IF NOT EXISTS rentals(
            id INTEGER PRIMARY KEY AUTOINCREMENT,customer_id INTEGER,vehicle_id INTEGER,
            booking_time TEXT,start_time TEXT,planned_days INTEGER DEFAULT 1,
            due_time TEXT,actual_return TEXT,price_per_day REAL DEFAULT 0,
            delivery_fee REAL DEFAULT 0,deposit REAL DEFAULT 0,rental_total REAL DEFAULT 0,
            extra_fee REAL DEFAULT 0,damage_fee REAL DEFAULT 0,fuel_fee REAL DEFAULT 0,
            discount REAL DEFAULT 0,final_total REAL DEFAULT 0,refund REAL DEFAULT 0,
            status TEXT DEFAULT 'Đang thuê',note TEXT,delivery_note TEXT);
        CREATE TABLE IF NOT EXISTS extensions(
            id INTEGER PRIMARY KEY AUTOINCREMENT,rental_id INTEGER,old_due TEXT,new_due TEXT,
            added_days INTEGER,added_amount REAL,created_at TEXT);
        CREATE TABLE IF NOT EXISTS expenses(
            id INTEGER PRIMARY KEY AUTOINCREMENT,expense_date TEXT,category TEXT,
            amount REAL DEFAULT 0,note TEXT);
        CREATE TABLE IF NOT EXISTS app_settings(key TEXT PRIMARY KEY,value TEXT);
        """)
        self.add_columns();self.c.commit()
    def add_columns(self):
        tables={"rentals":{"delivery_fee":"REAL DEFAULT 0","delivery_note":"TEXT","refund":"REAL DEFAULT 0"},"customers":{"created_at":"TEXT"}}
        for table,cols in tables.items():
            exists={r["name"] for r in self.c.execute(f"PRAGMA table_info({table})")}
            for name,definition in cols.items():
                if name not in exists:self.c.execute(f"ALTER TABLE {table} ADD COLUMN {name} {definition}")
        self.c.commit()
    def migrate_money(self):
        marker=self.one("SELECT value FROM app_settings WHERE key='money_unit'")
        if marker:return
        self.c.execute("UPDATE vehicles SET default_price=COALESCE(default_price,0)*1000")
        for col in ["price_per_day","delivery_fee","deposit","rental_total","extra_fee","damage_fee","fuel_fee","discount","final_total","refund"]:
            self.c.execute(f"UPDATE rentals SET {col}=COALESCE({col},0)*1000")
        self.c.execute("UPDATE extensions SET added_amount=COALESCE(added_amount,0)*1000")
        self.c.execute("UPDATE expenses SET amount=COALESCE(amount,0)*1000")
        self.c.execute("INSERT OR REPLACE INTO app_settings(key,value) VALUES('money_unit','VND')")
        self.c.commit()
    def all(self,sql,p=()):return self.c.execute(sql,p).fetchall()
    def one(self,sql,p=()):return self.c.execute(sql,p).fetchone()
    def run(self,sql,p=()):
        cur=self.c.execute(sql,p);self.c.commit();return cur.lastrowid
    def close(self):self.c.close()

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("QUẢN LÝ CHO THUÊ XE MÁY V3")
        self.geometry("1350x820");self.minsize(1100,700);self.configure(bg="#eef2f7")
        self.db=DB();self.selected_rental=None;self.vehicle_map={}
        self.style();self.interface();self.refresh_all()
        self.protocol("WM_DELETE_WINDOW",self.close_app);self.after(30000,self.auto_refresh)

    def style(self):
        s=ttk.Style(self);s.theme_use("clam")
        s.configure("Treeview",rowheight=29,font=("Segoe UI",9))
        s.configure("Treeview.Heading",font=("Segoe UI",9,"bold"))
        s.configure("TButton",padding=6)
        s.configure("TNotebook.Tab",font=("Segoe UI",9,"bold"),padding=(12,7))

    def interface(self):
        header=tk.Frame(self,bg="#172033",height=60);header.pack(fill="x")
        tk.Label(header,text="🏍  QUẢN LÝ CHO THUÊ XE MÁY V3",bg="#172033",fg="white",font=("Segoe UI",17,"bold")).pack(side="left",padx=18,pady=12)
        self.clock=tk.Label(header,bg="#172033",fg="#dbeafe",font=("Segoe UI",10));self.clock.pack(side="right",padx=18)
        self.update_clock()
        self.nb=ttk.Notebook(self);self.nb.pack(fill="both",expand=True,padx=8,pady=8)
        self.tab_dashboard=tk.Frame(self.nb,bg="#eef2f7")
        self.tab_rental=tk.Frame(self.nb,bg="#eef2f7")
        self.tab_vehicle=tk.Frame(self.nb,bg="#eef2f7")
        self.tab_expense=tk.Frame(self.nb,bg="#eef2f7")
        self.tab_report=tk.Frame(self.nb,bg="#eef2f7")
        for tab,text in [(self.tab_dashboard,"TỔNG QUAN"),(self.tab_rental,"GIAO / TRẢ XE"),(self.tab_vehicle,"XE"),(self.tab_expense,"CHI PHÍ"),(self.tab_report,"DOANH THU / EXCEL")]:
            self.nb.add(tab,text=text)
        self.build_dashboard();self.build_rental();self.build_vehicle();self.build_expense();self.build_report()

    def update_clock(self):
        self.clock.config(text=now().strftime("%d/%m/%Y  %H:%M:%S"));self.after(1000,self.update_clock)

    def field(self,parent,row,label,width=30):
        tk.Label(parent,text=label,bg="#eef2f7").grid(row=row,column=0,sticky="w",padx=8,pady=4)
        e=tk.Entry(parent,width=width);e.grid(row=row,column=1,padx=8,pady=4);return e

    def money_field(self,parent,row,label,default="0"):
        e=self.field(parent,row,label);e.insert(0,default)
        tk.Label(parent,text="VNĐ",bg="#eef2f7",fg="#b45309").grid(row=row,column=2,padx=2)
        return e

    def card(self,parent,title,value,col):
        f=tk.Frame(parent,bg="white",bd=1,relief="solid");f.grid(row=0,column=col,sticky="nsew",padx=5,pady=5)
        tk.Label(f,text=title,bg="white",fg="#64748b").pack(anchor="w",padx=12,pady=(10,2))
        l=tk.Label(f,text=value,bg="white",fg="#0f172a",font=("Segoe UI",16,"bold"));l.pack(anchor="w",padx=12,pady=(0,12));return l

    def build_dashboard(self):
        for i in range(4):self.tab_dashboard.grid_columnconfigure(i,weight=1)
        self.card_active=self.card(self.tab_dashboard,"XE ĐANG THUÊ","0",0)
        self.card_overdue=self.card(self.tab_dashboard,"ĐƠN QUÁ HẠN","0",1)
        self.card_free=self.card(self.tab_dashboard,"XE ĐANG TRỐNG","0",2)
        self.card_revenue=self.card(self.tab_dashboard,"DOANH THU THÁNG","0 đ",3)
        box=tk.LabelFrame(self.tab_dashboard,text="ĐƠN ĐANG THUÊ",bg="#eef2f7",font=("Segoe UI",10,"bold"))
        box.grid(row=1,column=0,columnspan=4,sticky="nsew",padx=5,pady=8);self.tab_dashboard.grid_rowconfigure(1,weight=1)
        cols=("xe","bien","khach","phone","giao","han","tong","status")
        self.dashboard_tree=ttk.Treeview(box,columns=cols,show="headings")
        heads=[("xe","Mã xe"),("bien","Biển số"),("khach","Khách hàng"),("phone","Điện thoại"),("giao","Ngày giờ giao"),("han","Hạn trả"),("tong","Tổng"),("status","Trạng thái")]
        for c,h in heads:self.dashboard_tree.heading(c,text=h);self.dashboard_tree.column(c,width=130)
        self.dashboard_tree.column("khach",width=170);self.dashboard_tree.tag_configure("overdue",background="#ffcccc",foreground="#a00000")
        self.dashboard_tree.pack(fill="both",expand=True,padx=5,pady=5)

    def build_rental(self):
        container=tk.Frame(self.tab_rental,bg="#eef2f7");container.pack(fill="both",expand=True)
        left=tk.LabelFrame(container,text="TẠO HỢP ĐỒNG + GIAO XE",bg="#eef2f7",font=("Segoe UI",10,"bold"));left.pack(side="left",fill="y",padx=(0,7))
        right=tk.LabelFrame(container,text="XE ĐANG THUÊ",bg="#eef2f7",font=("Segoe UI",10,"bold"));right.pack(side="left",fill="both",expand=True)
        self.rental_entries={};row=0
        for label,key in [("Tên khách hàng","name"),("Điện thoại","phone"),("CCCD / Hộ chiếu","id_number"),("Địa chỉ","address")]:
            self.rental_entries[key]=self.field(left,row,label);row+=1
        tk.Label(left,text="Xe",bg="#eef2f7").grid(row=row,column=0,sticky="w",padx=8,pady=4)
        self.vehicle_combo=ttk.Combobox(left,width=28,state="readonly");self.vehicle_combo.grid(row=row,column=1,padx=8,pady=4);self.vehicle_combo.bind("<<ComboboxSelected>>",self.vehicle_selected);row+=1
        self.price_entry=self.money_field(left,row,"Giá thuê / 24 giờ");row+=1
        tk.Label(left,text="Số ngày / 24 giờ",bg="#eef2f7").grid(row=row,column=0,sticky="w",padx=8,pady=4)
        self.days_spin=tk.Spinbox(left,from_=1,to=365,width=28);self.days_spin.delete(0,"end");self.days_spin.insert(0,"1");self.days_spin.grid(row=row,column=1,padx=8,pady=4);row+=1
        self.deposit_entry=self.money_field(left,row,"Tiền cọc");row+=1
        tk.Label(left,text="Ngày giao xe",bg="#eef2f7").grid(row=row,column=0,sticky="w",padx=8,pady=4)
        self.start_date=DateEntry(left,width=27,date_pattern="dd/mm/yyyy",locale="vi_VN");self.start_date.set_date(now());self.start_date.grid(row=row,column=1,padx=8,pady=4);row+=1
        tk.Label(left,text="Giờ giao xe",bg="#eef2f7").grid(row=row,column=0,sticky="w",padx=8,pady=4)
        self.start_time=ttk.Combobox(left,width=28,values=hours_list(),state="readonly");self.start_time.set(now().strftime("%H:%M"));self.start_time.grid(row=row,column=1,padx=8,pady=4);row+=1
        tk.Label(left,text="Ghi chú",bg="#eef2f7").grid(row=row,column=0,sticky="nw",padx=8,pady=4)
        self.note_text=tk.Text(left,width=30,height=4);self.note_text.grid(row=row,column=1,padx=8,pady=4);row+=1
        tk.Button(left,text="🚚 TẠO HỢP ĐỒNG + GIAO XE",bg="#16a34a",fg="white",font=("Segoe UI",10,"bold"),relief="flat",command=self.create_rental).grid(row=row,column=0,columnspan=3,sticky="ew",padx=8,pady=10);row+=1
        ttk.Button(left,text="🧾 Xuất phiếu hình ảnh",command=self.export_selected_receipt).grid(row=row,column=0,columnspan=3,sticky="ew",padx=8,pady=3);row+=1
        ttk.Button(left,text="↻ Xóa form",command=self.clear_rental_form).grid(row=row,column=0,columnspan=3,sticky="ew",padx=8,pady=3)
        toolbar=tk.Frame(right,bg="#eef2f7");toolbar.pack(fill="x",padx=5,pady=5)
        for text,command in [("🔄 Làm mới",self.refresh_rentals),("⏱ Gia hạn",self.extend_selected),("↩ Trả xe",self.return_selected),("🧾 Phiếu",self.export_selected_receipt)]:
            ttk.Button(toolbar,text=text,command=command).pack(side="left",padx=3)
        cols=("id","xe","bien","khach","phone","start","due","price","deposit","status")
        self.rental_tree=ttk.Treeview(right,columns=cols,show="headings")
        heads={"id":"Đơn","xe":"Mã xe","bien":"Biển số","khach":"Khách hàng","phone":"Điện thoại","start":"Ngày giờ giao","due":"Hạn trả","price":"Giá/24h","deposit":"Cọc","status":"Trạng thái"}
        widths={"id":50,"xe":70,"bien":90,"khach":150,"phone":105,"start":135,"due":135,"price":100,"deposit":100,"status":105}
        for c in cols:self.rental_tree.heading(c,text=heads[c]);self.rental_tree.column(c,width=widths[c],anchor="center")
        self.rental_tree.column("khach",anchor="w");self.rental_tree.tag_configure("overdue",background="#ffb3b3",foreground="#990000",font=("Segoe UI",9,"bold"));self.rental_tree.tag_configure("normal",background="#dcfce7")
        self.rental_tree.pack(fill="both",expand=True,padx=5,pady=5);self.rental_tree.bind("<<TreeviewSelect>>",self.rental_selected)

    def get_start_datetime(self):
        return datetime.strptime(f"{self.start_date.get_date().strftime('%d/%m/%Y')} {self.start_time.get()}",FMT)

    def build_vehicle(self):
        form=tk.LabelFrame(self.tab_vehicle,text="THÊM XE",bg="#eef2f7",font=("Segoe UI",10,"bold"));form.pack(fill="x",padx=5,pady=5)
        self.vehicle_inputs={}
        for i,(label,key) in enumerate([("Mã xe","code"),("Biển số","plate"),("Dòng xe","model"),("Màu","color"),("Giá / 24h","price")]):
            tk.Label(form,text=label,bg="#eef2f7").grid(row=0,column=i*2,padx=5,pady=7)
            e=tk.Entry(form,width=15);e.grid(row=0,column=i*2+1,padx=5);self.vehicle_inputs[key]=e
        tk.Label(form,text="VNĐ",bg="#eef2f7",fg="#b45309").grid(row=0,column=12)
        ttk.Button(form,text="➕ Thêm xe",command=self.add_vehicle).grid(row=1,column=0,columnspan=2,pady=6)
        ttk.Button(form,text="🗑 Xóa xe",command=self.delete_vehicle).grid(row=1,column=2,columnspan=2,pady=6)
        box=tk.Frame(self.tab_vehicle,bg="#eef2f7");box.pack(fill="both",expand=True,padx=5,pady=5)
        cols=("code","plate","model","color","price","status","note");self.vehicle_tree=ttk.Treeview(box,columns=cols,show="headings")
        for c,h in zip(cols,["Mã xe","Biển số","Dòng xe","Màu","Giá / 24h","Trạng thái","Ghi chú"]):
            self.vehicle_tree.heading(c,text=h);self.vehicle_tree.column(c,width=140)
        self.vehicle_tree.tag_configure("rented",background="#fef3c7");self.vehicle_tree.pack(fill="both",expand=True)

    def refresh_vehicles(self):
        for i in self.vehicle_tree.get_children():self.vehicle_tree.delete(i)
        for r in self.db.all("SELECT * FROM vehicles ORDER BY code"):
            self.vehicle_tree.insert("","end",values=(r["code"],r["plate"],r["model"],r["color"] or "",money(r["default_price"]),r["status"],r["note"] or ""),tags=("rented" if r["status"]!="Trống" else "",))

    def refresh_vehicle_combo(self):
        rows=self.db.all("SELECT * FROM vehicles WHERE status='Trống' ORDER BY code");self.vehicle_map={};values=[]
        for r in rows:
            text=f"{r['code']} | {r['plate']} | {r['model']}";values.append(text);self.vehicle_map[text]=r
        self.vehicle_combo["values"]=values

    def vehicle_selected(self,event=None):
        r=self.vehicle_map.get(self.vehicle_combo.get())
        if r:
            self.price_entry.delete(0,"end");self.price_entry.insert(0,str(int(num(r["default_price"]))))

    def add_vehicle(self):
        try:
            code=self.vehicle_inputs["code"].get().strip();plate=self.vehicle_inputs["plate"].get().strip();model=self.vehicle_inputs["model"].get().strip();color=self.vehicle_inputs["color"].get().strip();price=money_input(self.vehicle_inputs["price"].get())
            if not code:raise ValueError("Chưa nhập mã xe.")
            if not plate:raise ValueError("Chưa nhập biển số.")
            if not model:raise ValueError("Chưa nhập dòng xe.")
            self.db.run("INSERT INTO vehicles(code,plate,model,color,default_price) VALUES(?,?,?,?,?)",(code,plate,model,color,price))
            for e in self.vehicle_inputs.values():e.delete(0,"end")
            self.refresh_all();messagebox.showinfo("Thành công","Đã thêm xe.")
        except sqlite3.IntegrityError:messagebox.showerror("Lỗi","Mã xe hoặc biển số đã tồn tại.")
        except Exception as e:messagebox.showerror("Lỗi",str(e))

    def delete_vehicle(self):
        selected=self.vehicle_tree.selection()
        if not selected:return messagebox.showwarning("Thông báo","Hãy chọn xe cần xóa.")
        code=self.vehicle_tree.item(selected[0],"values")[0];v=self.db.one("SELECT * FROM vehicles WHERE code=?",(code,))
        if not v:return
        if v["status"]!="Trống":return messagebox.showwarning("Không thể xóa","Xe đang được thuê.")
        if messagebox.askyesno("Xác nhận",f"Bạn có chắc muốn xóa xe {code}?"):
            self.db.run("DELETE FROM vehicles WHERE id=?",(v["id"],));self.refresh_all()

    def create_rental(self):
        try:
            name=self.rental_entries["name"].get().strip();phone=self.rental_entries["phone"].get().strip();id_number=self.rental_entries["id_number"].get().strip();address=self.rental_entries["address"].get().strip()
            if not name:raise ValueError("Chưa nhập tên khách hàng.")
            if not phone:raise ValueError("Chưa nhập số điện thoại.")
            vehicle=self.vehicle_map.get(self.vehicle_combo.get())
            if not vehicle:raise ValueError("Chưa chọn xe.")
            price=money_input(self.price_entry.get());days=int(self.days_spin.get());deposit=money_input(self.deposit_entry.get());start=self.get_start_datetime()
            if days<=0:raise ValueError("Số ngày phải lớn hơn 0.")
            due=start+timedelta(hours=24*days);note=self.note_text.get("1.0","end").strip()
            customer=self.db.one("SELECT * FROM customers WHERE phone=? ORDER BY id DESC LIMIT 1",(phone,))
            if customer:
                customer_id=customer["id"];self.db.run("UPDATE customers SET name=?,id_number=?,address=?,note=? WHERE id=?",(name,id_number,address,note,customer_id))
            else:
                customer_id=self.db.run("INSERT INTO customers(name,phone,id_number,address,note,created_at) VALUES(?,?,?,?,?,?)",(name,phone,id_number,address,note,now().strftime(FMT)))
            total=price*days
            rental_id=self.db.run("""INSERT INTO rentals(customer_id,vehicle_id,booking_time,start_time,planned_days,due_time,price_per_day,deposit,rental_total,final_total,status,note,delivery_note) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""",(customer_id,vehicle["id"],now().strftime(FMT),start.strftime(FMT),days,due.strftime(FMT),price,deposit,total,total,"Đang thuê",note,note))
            self.db.run("UPDATE vehicles SET status='Đang thuê' WHERE id=?",(vehicle["id"],));self.selected_rental=rental_id
            receipt=self.create_receipt_image(rental_id);self.refresh_all()
            messagebox.showinfo("GIAO XE THÀNH CÔNG",f"Đơn #{rental_id}\n\nXe: {vehicle['code']} - {vehicle['plate']}\nKhách: {name}\n\nGiao xe:\n{start.strftime(FMT)}\n\nHạn trả:\n{due.strftime(FMT)}\n\nTiền thuê:\n{money(total)}\n\nTiền cọc:\n{money(deposit)}\n\nPhiếu:\n{receipt}")
            self.clear_rental_form()
        except Exception as e:messagebox.showerror("Không thể giao xe",str(e))

    def active_rentals(self):
        return self.db.all("""SELECT r.*,c.name customer_name,c.phone customer_phone,c.id_number customer_id_number,c.address customer_address,v.code vehicle_code,v.plate vehicle_plate,v.model vehicle_model,v.color vehicle_color FROM rentals r LEFT JOIN customers c ON c.id=r.customer_id LEFT JOIN vehicles v ON v.id=r.vehicle_id WHERE r.status IN('Đang thuê','Gia hạn') ORDER BY r.due_time""")

    def overdue(self,r):
        try:return dt(r["due_time"])<now() and r["status"] in ("Đang thuê","Gia hạn")
        except:return False

    def refresh_rentals(self):
        for i in self.rental_tree.get_children():self.rental_tree.delete(i)
        for r in self.active_rentals():
            late=self.overdue(r)
            self.rental_tree.insert("","end",iid=str(r["id"]),values=(r["id"],r["vehicle_code"],r["vehicle_plate"],r["customer_name"] or "",r["customer_phone"] or "",r["start_time"],r["due_time"],money(r["price_per_day"]),money(r["deposit"]),"⚠ QUÁ HẠN" if late else r["status"]),tags=("overdue" if late else "normal",))
        self.refresh_vehicle_combo()

    def rental_selected(self,event=None):
        s=self.rental_tree.selection();self.selected_rental=int(s[0]) if s else None

    def extend_selected(self):
        if not self.selected_rental:return messagebox.showwarning("Thông báo","Hãy chọn đơn cần gia hạn.")
        r=self.db.one("SELECT * FROM rentals WHERE id=?",(self.selected_rental,))
        if not r or r["status"] not in ("Đang thuê","Gia hạn"):return messagebox.showwarning("Không thể gia hạn","Đơn này đã trả xe.")
        w=tk.Toplevel(self);w.title(f"GIA HẠN ĐƠN #{r['id']}");w.transient(self);w.grab_set()
        tk.Label(w,text=f"Hạn hiện tại: {r['due_time']}",font=("Segoe UI",10,"bold")).pack(padx=15,pady=10)
        tk.Label(w,text="Số ngày gia hạn").pack()
        days=tk.Spinbox(w,from_=1,to=365,width=10);days.pack(pady=5)
        tk.Label(w,text="Giá / 24 giờ (VNĐ)").pack()
        price=tk.Entry(w,width=20);price.insert(0,str(int(num(r["price_per_day"]))));price.pack(pady=5)
        def save():
            try:
                d=int(days.get());p=money_input(price.get())
                if d<=0:raise ValueError("Số ngày gia hạn không hợp lệ.")
                old=dt(r["due_time"]);new=old+timedelta(hours=24*d);added=p*d
                self.db.run("INSERT INTO extensions(rental_id,old_due,new_due,added_days,added_amount,created_at) VALUES(?,?,?,?,?,?)",(r["id"],r["due_time"],new.strftime(FMT),d,added,now().strftime(FMT)))
                self.db.run("UPDATE rentals SET planned_days=planned_days+?,due_time=?,rental_total=rental_total+?,final_total=final_total+?,status='Gia hạn' WHERE id=?",(d,new.strftime(FMT),added,added,r["id"]))
                w.destroy();self.refresh_all();messagebox.showinfo("Đã gia hạn",f"Hạn mới:\n{new.strftime(FMT)}\n\nTiền gia hạn:\n{money(added)}")
            except Exception as e:messagebox.showerror("Lỗi",str(e))
        ttk.Button(w,text="LƯU GIA HẠN",command=save).pack(fill="x",padx=15,pady=15)

    def return_selected(self):
        if not self.selected_rental:return messagebox.showwarning("Thông báo","Hãy chọn đơn cần trả.")
        r=self.db.one("""SELECT r.*,c.name customer_name,c.phone,v.code,v.plate,v.model FROM rentals r LEFT JOIN customers c ON c.id=r.customer_id LEFT JOIN vehicles v ON v.id=r.vehicle_id WHERE r.id=?""",(self.selected_rental,))
        if not r:return
        w=tk.Toplevel(self);w.title(f"TRẢ XE - ĐƠN #{r['id']}");w.transient(self);w.grab_set();entries={}
        fields=[("Ngày giờ trả xe",now().strftime(FMT)),("Hoàn tiền khách","0"),("Phí hư hỏng","0"),("Phí xăng","0"),("Giảm giá","0")]
        for i,(label,value) in enumerate(fields):
            tk.Label(w,text=label).grid(row=i,column=0,sticky="w",padx=10,pady=5)
            e=tk.Entry(w,width=28);e.insert(0,value);e.grid(row=i,column=1,padx=10,pady=5);entries[label]=e
        tk.Label(w,text="Hoàn tiền khách sẽ được trừ khỏi tổng tiền.",fg="#b45309").grid(row=6,column=1,sticky="w")
        def confirm():
            try:
                return_dt=dt(entries["Ngày giờ trả xe"].get());start_dt=dt(r["start_time"])
                if return_dt<start_dt:raise ValueError("Ngày giờ trả không được trước ngày giao.")
                refund=money_input(entries["Hoàn tiền khách"].get());damage=money_input(entries["Phí hư hỏng"].get());fuel=money_input(entries["Phí xăng"].get());discount=money_input(entries["Giảm giá"].get())
                total=num(r["rental_total"])+damage+fuel-discount-refund
                if total<0:raise ValueError("Tổng tiền không thể nhỏ hơn 0.")
                self.db.run("UPDATE rentals SET actual_return=?,refund=?,damage_fee=?,fuel_fee=?,discount=?,final_total=?,status='Đã trả' WHERE id=?",(return_dt.strftime(FMT),refund,damage,fuel,discount,total,r["id"]))
                self.db.run("UPDATE vehicles SET status='Trống' WHERE id=?",(r["vehicle_id"],));w.destroy();self.refresh_all()
                messagebox.showinfo("TRẢ XE THÀNH CÔNG",f"Đơn #{r['id']}\n\nXe: {r['code']} - {r['plate']}\nNgày giờ trả: {return_dt.strftime(FMT)}\n\nHoàn tiền khách: {money(refund)}\nPhí hư hỏng: {money(damage)}\nPhí xăng: {money(fuel)}\nGiảm giá: {money(discount)}\n\nTỔNG TIỀN: {money(total)}\n\nXe đã chuyển về TRỐNG.")
            except Exception as e:messagebox.showerror("Lỗi trả xe",str(e))
        ttk.Button(w,text="✓ XÁC NHẬN TRẢ XE",command=confirm).grid(row=7,column=0,columnspan=2,sticky="ew",padx=10,pady=12)

    def rental_full(self,rid):
        return self.db.one("""SELECT r.*,c.name customer_name,c.phone customer_phone,c.id_number customer_id_number,c.address customer_address,v.code vehicle_code,v.plate vehicle_plate,v.model vehicle_model,v.color vehicle_color FROM rentals r LEFT JOIN customers c ON c.id=r.customer_id LEFT JOIN vehicles v ON v.id=r.vehicle_id WHERE r.id=?""",(rid,))

    def create_receipt_image(self,rid):
        r=self.rental_full(rid)
        if not r:raise ValueError("Không tìm thấy đơn.")
        if not PIL_OK:raise ValueError("Máy chưa cài Pillow.\n\npip install pillow")
        start_dt=dt(r["start_time"]);date_str=start_dt.strftime("%d-%m-%Y");png=os.path.join(PHIEU_DIR,f"PHIEU GIAO XE SỐ {rid}_{date_str}.png")
        width,height=1100,1850;im=Image.new("RGB",(width,height),"white");d=ImageDraw.Draw(im)
        normal=["C:/Windows/Fonts/arial.ttf","C:/Windows/Fonts/segoeui.ttf","C:/Windows/Fonts/tahoma.ttf"];bold=["C:/Windows/Fonts/arialbd.ttf","C:/Windows/Fonts/segoeuib.ttf","C:/Windows/Fonts/tahomabd.ttf"]
        nf=next((x for x in normal if os.path.exists(x)),None);bf=next((x for x in bold if os.path.exists(x)),nf)
        if nf:title=ImageFont.truetype(bf,34);head=ImageFont.truetype(bf,23);font=ImageFont.truetype(nf,20);small=ImageFont.truetype(nf,17)
        else:title=head=font=small=ImageFont.load_default()
        x,y=60,35;d.text((x,y),"PHIẾU GIAO XE MÁY",font=title,fill="#172033");y+=60;d.text((x,y),f"PHIẾU SỐ {rid} - {date_str}",font=head,fill="#334155");y+=45;d.line((x,y,width-x,y),fill="#94a3b8",width=2);y+=30
        sections=[("KHÁCH HÀNG",[("Tên",r["customer_name"] or ""),("Điện thoại",r["customer_phone"] or ""),("CCCD/Hộ chiếu",r["customer_id_number"] or ""),("Địa chỉ",r["customer_address"] or "")]),("XE",[("Mã xe",r["vehicle_code"]),("Biển số",r["vehicle_plate"]),("Dòng xe",r["vehicle_model"]),("Màu",r["vehicle_color"] or "")]),("THỜI GIAN",[("Ngày giờ giao",r["start_time"]),("Ngày giờ hết hạn",r["due_time"]),("Số ngày",f"{r['planned_days']} x 24 giờ")])]
        for title_text,lines in sections:
            d.text((x,y),title_text,font=head,fill="#1d4ed8");y+=38
            for label,value in lines:d.text((x+15,y),label,font=small,fill="#64748b");d.text((x+270,y),str(value),font=font,fill="#111827");y+=38
            y+=10
        d.line((x,y,width-x,y),fill="#cbd5e1",width=2);y+=25;d.text((x,y),"THANH TOÁN",font=head,fill="#15803d");y+=42
        for label,value in [("Giá / 24 giờ",money(r["price_per_day"])),("Tiền cọc",money(r["deposit"])),("TỔNG TIỀN",money(r["rental_total"]))]:
            d.text((x+15,y),label,font=font,fill="#111827");d.text((x+650,y),value,font=font,fill="#111827");y+=40
        y+=15;d.line((x,y,width-x,y),fill="#cbd5e1",width=2);y+=25;d.text((x,y),"GHI CHÚ",font=head,fill="#334155");y+=38
        note=r["delivery_note"] or ""
        if note:
            for i in range(0,len(note),75):d.text((x+15,y),note[i:i+75],font=font,fill="#111827");y+=30
        else:d.text((x+15,y),"Không có",font=font,fill="#64748b");y+=30
        y+=20;d.text((x,y),"LƯU Ý",font=head,fill="#b91c1c");y+=38
        for text in ["• Thời gian thuê tính theo 24 giờ kể từ lúc giao xe.","• Đơn quá hạn sẽ tự động hiển thị màu đỏ."]:d.text((x+15,y),text,font=small,fill="#334155");y+=30
        y+=35;d.text((x+20,y),"Nhân viên: __________________________",font=font,fill="#111827");d.text((x+570,y),"Khách hàng: __________________",font=font,fill="#111827");y+=60;d.line((x,y,width-x,y),fill="#94a3b8",width=2);y+=25
        if os.path.exists(QR_FILE):
            try:
                qr=Image.open(QR_FILE).convert("RGB");qr_size=280;qr.thumbnail((qr_size,qr_size),Image.Resampling.LANCZOS);qr_x=(width-qr.width)//2;qr_y=y+30;text="MÃ QR THANH TOÁN";bbox=d.textbbox((0,0),text,font=head);d.text(((width-(bbox[2]-bbox[0]))//2,y),text,font=head,fill="#1d4ed8");im.paste(qr,(qr_x,qr_y));text2="Quét mã QR để thanh toán";bbox2=d.textbbox((0,0),text2,font=small);d.text(((width-(bbox2[2]-bbox2[0]))//2,qr_y+qr.height+15),text2,font=small,fill="#475569")
            except:d.text((x,y),"QR_THANH_TOAN.png không đọc được.",font=font,fill="#b91c1c")
        else:
            text="CHƯA CÓ QR THANH TOÁN";bbox=d.textbbox((0,0),text,font=head);d.text(((width-(bbox[2]-bbox[0]))//2,y+20),text,font=head,fill="#b91c1c")
            text2="Đặt file QR_THANH_TOAN.png cạnh file chương trình.";bbox2=d.textbbox((0,0),text2,font=small);d.text(((width-(bbox2[2]-bbox2[0]))//2,y+65),text2,font=small,fill="#64748b")
        im.save(png);return png

    def export_selected_receipt(self):
        if not self.selected_rental:return messagebox.showwarning("Thông báo","Hãy chọn một đơn.")
        try:
            path=self.create_receipt_image(self.selected_rental);messagebox.showinfo("Đã tạo phiếu",f"Phiếu giao xe đã được tạo:\n\n{path}")
        except Exception as e:messagebox.showerror("Lỗi",str(e))

    def clear_rental_form(self):
        for e in self.rental_entries.values():e.delete(0,"end")
        self.vehicle_combo.set("");self.price_entry.delete(0,"end");self.days_spin.delete(0,"end");self.days_spin.insert(0,"1");self.deposit_entry.delete(0,"end");self.deposit_entry.insert(0,"0");self.start_date.set_date(now());self.start_time.set(now().strftime("%H:%M"));self.note_text.delete("1.0","end")

    def build_expense(self):
        form=tk.LabelFrame(self.tab_expense,text="NHẬP CHI PHÍ",bg="#eef2f7",font=("Segoe UI",10,"bold"));form.pack(fill="x",padx=5,pady=5)
        tk.Label(form,text="Ngày",bg="#eef2f7").grid(row=0,column=0,padx=5,pady=5);self.exp_date=tk.Entry(form,width=15);self.exp_date.insert(0,now().strftime("%d/%m/%Y"));self.exp_date.grid(row=0,column=1)
        tk.Label(form,text="Loại",bg="#eef2f7").grid(row=0,column=2);self.exp_category=tk.Entry(form,width=18);self.exp_category.grid(row=0,column=3)
        tk.Label(form,text="Số tiền",bg="#eef2f7").grid(row=0,column=4);self.exp_amount=tk.Entry(form,width=15);self.exp_amount.grid(row=0,column=5);tk.Label(form,text="VNĐ",bg="#eef2f7",fg="#b45309").grid(row=0,column=6)
        tk.Label(form,text="Ghi chú",bg="#eef2f7").grid(row=1,column=0);self.exp_note=tk.Entry(form,width=60);self.exp_note.grid(row=1,column=1,columnspan=5,pady=5);ttk.Button(form,text="➕ Lưu",command=self.add_expense).grid(row=1,column=6)
        cols=("date","category","amount","note");self.expense_tree=ttk.Treeview(self.tab_expense,columns=cols,show="headings")
        for c,h in zip(cols,["Ngày","Loại","Số tiền","Ghi chú"]):self.expense_tree.heading(c,text=h);self.expense_tree.column(c,width=200)
        self.expense_tree.pack(fill="both",expand=True,padx=5,pady=5)

    def add_expense(self):
        try:
            date=self.exp_date.get().strip();datetime.strptime(date,"%d/%m/%Y");category=self.exp_category.get().strip() or "Khác";amount=money_input(self.exp_amount.get());note=self.exp_note.get().strip()
            self.db.run("INSERT INTO expenses(expense_date,category,amount,note) VALUES(?,?,?,?)",(date,category,amount,note));self.exp_amount.delete(0,"end");self.exp_note.delete(0,"end");self.refresh_expenses();self.refresh_report()
        except Exception as e:messagebox.showerror("Lỗi",str(e))

    def refresh_expenses(self):
        for i in self.expense_tree.get_children():self.expense_tree.delete(i)
        for r in self.db.all("SELECT * FROM expenses ORDER BY id DESC"):self.expense_tree.insert("","end",values=(r["expense_date"],r["category"],money(r["amount"]),r["note"] or ""))

    def build_report(self):
        toolbar=tk.Frame(self.tab_report,bg="#eef2f7");toolbar.pack(fill="x",padx=5,pady=7);tk.Label(toolbar,text="Tháng",bg="#eef2f7").pack(side="left")
        self.report_month=ttk.Combobox(toolbar,values=[f"{i:02d}" for i in range(1,13)],width=5,state="readonly");self.report_month.set(f"{now().month:02d}");self.report_month.pack(side="left",padx=5)
        tk.Label(toolbar,text="Năm",bg="#eef2f7").pack(side="left");self.report_year=tk.Entry(toolbar,width=8);self.report_year.insert(0,str(now().year));self.report_year.pack(side="left",padx=5)
        ttk.Button(toolbar,text="📊 Xem",command=self.refresh_report).pack(side="left",padx=5);ttk.Button(toolbar,text="📗 Xuất Excel",command=self.export_excel).pack(side="left",padx=5)
        self.report_text=tk.Text(self.tab_report,bg="white",font=("Consolas",10));self.report_text.pack(fill="both",expand=True,padx=5,pady=5)

    def report_data(self,month,year):
        rows=self.db.all("""SELECT r.*,c.name customer_name,c.phone,v.code,v.plate,v.model FROM rentals r LEFT JOIN customers c ON c.id=r.customer_id LEFT JOIN vehicles v ON v.id=r.vehicle_id WHERE r.status='Đã trả' AND substr(r.actual_return,7,4)=? AND substr(r.actual_return,4,2)=? ORDER BY r.actual_return""",(str(year),f"{month:02d}"))
        expenses=self.db.all("SELECT * FROM expenses WHERE substr(expense_date,7,4)=? AND substr(expense_date,4,2)=? ORDER BY expense_date",(str(year),f"{month:02d}"))
        return rows,expenses

    def refresh_report(self):
        try:month=int(self.report_month.get());year=int(self.report_year.get())
        except:return
        rows,expenses=self.report_data(month,year);revenue=sum(num(r["final_total"]) for r in rows);cost=sum(num(e["amount"]) for e in expenses);profit=revenue-cost
        self.report_text.delete("1.0","end");self.report_text.insert("end",f"DOANH THU THÁNG {month:02d}/{year}\n"+"="*90+"\n\n");self.report_text.insert("end",f"Số đơn đã trả : {len(rows)}\nDOANH THU     : {money(revenue)}\nCHI PHÍ       : {money(cost)}\nLỢI NHUẬN     : {money(profit)}\n\n");self.report_text.insert("end","CHI TIẾT ĐƠN\n"+"-"*90+"\n")
        for r in rows:self.report_text.insert("end",f"#{r['id']} | {r['actual_return']} | {r['code']} | {r['plate']} | {r['customer_name'] or ''} | {money(r['final_total'])}\n")
        self.report_text.insert("end","\nCHI PHÍ\n"+"-"*90+"\n")
        for e in expenses:self.report_text.insert("end",f"{e['expense_date']} | {e['category']} | {money(e['amount'])} | {e['note'] or ''}\n")

    def export_excel(self):
        if not XLSX_OK:return messagebox.showerror("Thiếu thư viện","Hãy chạy:\npip install openpyxl")
        try:
            month=int(self.report_month.get());year=int(self.report_year.get());rows,expenses=self.report_data(month,year);wb=Workbook();ws=wb.active;ws.title="Doanh thu"
            headers=["Đơn","Ngày trả","Mã xe","Biển số","Dòng xe","Khách hàng","Điện thoại","Ngày giờ giao","Hạn trả","Giá / 24h (VNĐ)","Cọc (VNĐ)","Hoàn tiền khách (VNĐ)","Phí hư hỏng (VNĐ)","Xăng (VNĐ)","Giảm giá (VNĐ)","Tổng tiền (VNĐ)","Trạng thái"];ws.append(headers)
            for c in ws[1]:c.font=Font(bold=True,color="FFFFFF");c.fill=PatternFill("solid",fgColor="1F4E78");c.alignment=Alignment(horizontal="center")
            for r in rows:ws.append([r["id"],r["actual_return"],r["code"],r["plate"],r["model"],r["customer_name"],r["phone"],r["start_time"],r["due_time"],r["price_per_day"],r["deposit"],r["refund"],r["damage_fee"],r["fuel_fee"],r["discount"],r["final_total"],r["status"]])
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if 10<=cell.column<=16:cell.number_format='#,##0" đ"'
            for col in ws.columns:
                length=max(len(str(c.value or "")) for c in col);ws.column_dimensions[get_column_letter(col[0].column)].width=min(max(length+2,12),30)
            ws.freeze_panes="A2";ex=wb.create_sheet("Chi phí");ex.append(["Ngày","Loại","Số tiền (VNĐ)","Ghi chú"])
            for c in ex[1]:c.font=Font(bold=True,color="FFFFFF");c.fill=PatternFill("solid",fgColor="7F6000")
            for e in expenses:ex.append([e["expense_date"],e["category"],e["amount"],e["note"] or ""])
            for row in ex.iter_rows(min_row=2):row[2].number_format='#,##0" đ"'
            sm=wb.create_sheet("Tổng hợp");revenue=sum(num(r["final_total"]) for r in rows);cost=sum(num(e["amount"]) for e in expenses);profit=revenue-cost
            sm.append(["BÁO CÁO DOANH THU",f"{month:02d}/{year}"]);sm.append(["Số đơn đã trả",len(rows)]);sm.append(["Doanh thu (VNĐ)",revenue]);sm.append(["Chi phí (VNĐ)",cost]);sm.append(["Lợi nhuận (VNĐ)",profit]);sm["A1"].font=Font(bold=True,size=15)
            for cell in ["B3","B4","B5"]:sm[cell].number_format='#,##0" đ"'
            sm.column_dimensions["A"].width=25;sm.column_dimensions["B"].width=25
            path=filedialog.asksaveasfilename(title="Lưu báo cáo Excel",defaultextension=".xlsx",initialfile=f"Doanh_thu_{year}_{month:02d}.xlsx",filetypes=[("Excel","*.xlsx")])
            if not path:return
            wb.save(path);messagebox.showinfo("Đã xuất Excel",f"Đã lưu:\n\n{path}")
        except Exception as e:messagebox.showerror("Lỗi xuất Excel",str(e))

    def refresh_dashboard(self):
        active=self.active_rentals();overdue=[r for r in active if self.overdue(r)];free=self.db.one("SELECT COUNT(*) total FROM vehicles WHERE status='Trống'")["total"]
        revenue=self.db.one("SELECT COALESCE(SUM(final_total),0) total FROM rentals WHERE status='Đã trả' AND substr(actual_return,7,4)=? AND substr(actual_return,4,2)=?",(str(now().year),f"{now().month:02d}"))["total"]
        self.card_active.config(text=str(len(active)));self.card_overdue.config(text=str(len(overdue)));self.card_free.config(text=str(free));self.card_revenue.config(text=money(revenue))
        for i in self.dashboard_tree.get_children():self.dashboard_tree.delete(i)
        for r in active:
            late=self.overdue(r);self.dashboard_tree.insert("","end",values=(r["vehicle_code"],r["vehicle_plate"],r["customer_name"] or "",r["customer_phone"] or "",r["start_time"],r["due_time"],money(r["rental_total"]),"⚠ QUÁ HẠN" if late else r["status"]),tags=("overdue" if late else "",))

    def refresh_all(self):
        self.refresh_vehicles();self.refresh_vehicle_combo();self.refresh_rentals();self.refresh_expenses();self.refresh_dashboard();self.refresh_report()

    def auto_refresh(self):
        self.refresh_rentals();self.refresh_dashboard();self.after(30000,self.auto_refresh)

    def close_app(self):
        try:self.db.close()
        except:pass
        self.destroy()

if __name__=="__main__":
    app=App()
    app.mainloop()
