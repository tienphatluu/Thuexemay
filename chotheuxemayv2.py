# -*- coding: utf-8 -*-

import os
import sqlite3
from datetime import datetime, timedelta

import streamlit as st

# =========================================================
# THƯ VIỆN TÙY CHỌN
# =========================================================

try:
    from PIL import Image, ImageDraw, ImageFont
    PIL_OK = True
except ImportError:
    PIL_OK = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    XLSX_OK = True
except ImportError:
    XLSX_OK = False


# =========================================================
# CẤU HÌNH
# =========================================================

BASE = os.path.dirname(os.path.abspath(__file__))

DB_FILE = os.path.join(BASE, "rental_v2.db")
PHIEU_DIR = os.path.join(BASE, "PHIEU_GIAO_XE")
QR_FILE = os.path.join(BASE, "QR_THANH_TOAN.png")

FMT = "%d/%m/%Y %H:%M"

os.makedirs(PHIEU_DIR, exist_ok=True)


# =========================================================
# PAGE CONFIG
# =========================================================

st.set_page_config(
    page_title="Quản lý cho thuê xe máy",
    page_icon="🏍️",
    layout="wide",
    initial_sidebar_state="expanded",
)


# =========================================================
# CSS GIAO DIỆN
# =========================================================

st.markdown(
    """
<style>

html, body, [class*="css"] {
    font-family: "Segoe UI", Arial, sans-serif;
}

/* Nền chính */
.stApp {
    background: #f1f5f9;
}

/* Header */
.main-header {
    background: linear-gradient(135deg, #172033, #243b64);
    padding: 20px 25px;
    border-radius: 14px;
    margin-bottom: 18px;
    color: white;
    box-shadow: 0 5px 18px rgba(15,23,42,.15);
}

.main-header h1 {
    margin: 0;
    color: white !important;
    font-size: 28px;
}

.main-header p {
    margin: 5px 0 0 0;
    color: #dbeafe !important;
}

/* Sidebar */
section[data-testid="stSidebar"] {
    background: #172033;
}

section[data-testid="stSidebar"] * {
    color: #f8fafc !important;
}

section[data-testid="stSidebar"] .stRadio label {
    color: #f8fafc !important;
}

/* Text */
label,
.stMarkdown,
.stTextInput label,
.stNumberInput label,
.stSelectbox label,
.stDateInput label,
.stTimeInput label,
.stTextArea label {
    color: #172033 !important;
    font-weight: 600 !important;
}

/* Input */
input,
textarea {
    color: #111827 !important;
    background: white !important;
    border: 1px solid #cbd5e1 !important;
    border-radius: 8px !important;
}

textarea {
    min-height: 70px !important;
}

div[data-baseweb="select"] > div {
    background: white !important;
    color: #111827 !important;
    border-color: #cbd5e1 !important;
}

/* Button */
.stButton > button {
    border-radius: 8px;
    border: 1px solid #cbd5e1;
    font-weight: 700;
    min-height: 42px;
    color: #172033 !important;
    background: white;
}

.stButton > button:hover {
    border-color: #2563eb;
    color: #2563eb !important;
}

/* Card */
.dashboard-card {
    background: white;
    border-radius: 13px;
    padding: 17px;
    border: 1px solid #e2e8f0;
    box-shadow: 0 3px 12px rgba(15,23,42,.06);
    min-height: 115px;
}

.dashboard-title {
    color: #64748b !important;
    font-size: 14px;
    font-weight: 600;
}

.dashboard-value {
    color: #0f172a !important;
    font-size: 27px;
    font-weight: 800;
    margin-top: 10px;
}

/* Section */
.section-box {
    background: white;
    padding: 18px;
    border-radius: 13px;
    border: 1px solid #e2e8f0;
    margin-bottom: 15px;
}

/* Title */
.section-title {
    color: #172033 !important;
    font-size: 18px;
    font-weight: 800;
    margin-bottom: 14px;
}

/* Status */
.status-rented {
    color: #b45309 !important;
    font-weight: 700;
}

.status-free {
    color: #15803d !important;
    font-weight: 700;
}

.status-overdue {
    color: #b91c1c !important;
    font-weight: 800;
}

/* Dataframe */
div[data-testid="stDataFrame"] {
    border-radius: 10px;
    overflow: hidden;
}

/* Tabs */
button[data-baseweb="tab"] {
    font-weight: 700 !important;
    color: #334155 !important;
}

button[data-baseweb="tab"][aria-selected="true"] {
    color: #1d4ed8 !important;
}

/* Expander */
.streamlit-expanderHeader {
    color: #172033 !important;
    font-weight: 700;
}

/* Alert */
.stAlert {
    border-radius: 10px;
}

/* Mobile */
@media (max-width: 900px) {
    .main-header h1 {
        font-size: 21px;
    }

    .dashboard-value {
        font-size: 22px;
    }
}

</style>
""",
    unsafe_allow_html=True,
)


# =========================================================
# HÀM CHUNG
# =========================================================

def now():
    return datetime.now()


def money(v):
    try:
        return f"{float(v):,.0f}".replace(",", ".") + " đ"
    except Exception:
        return "0 đ"


def num(v):
    try:
        return float(v)
    except Exception:
        return 0


def money_input(v):
    s = (
        str(v)
        .strip()
        .replace(" ", "")
        .replace("đ", "")
        .replace("Đ", "")
        .replace(".", "")
        .replace(",", "")
    )

    if not s:
        return 0

    try:
        n = float(s)
    except Exception:
        raise ValueError("Số tiền không hợp lệ.")

    if n < 0:
        raise ValueError("Số tiền không được âm.")

    return n


def dt(v):
    return datetime.strptime(str(v).strip(), FMT)


def fmt_date(d):
    return d.strftime("%d/%m/%Y")


def is_overdue(r):
    try:
        return (
            dt(r["due_time"]) < now()
            and r["status"] in ("Đang thuê", "Gia hạn")
        )
    except Exception:
        return False


# =========================================================
# DATABASE
# =========================================================

class DB:

    def __init__(self):
        self.c = sqlite3.connect(
            DB_FILE,
            check_same_thread=False
        )
        self.c.row_factory = sqlite3.Row

        self.create()
        self.migrate_money()

    def create(self):

        self.c.executescript(
            """
            CREATE TABLE IF NOT EXISTS vehicles(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT UNIQUE NOT NULL,
                plate TEXT UNIQUE NOT NULL,
                model TEXT,
                color TEXT,
                default_price REAL DEFAULT 0,
                status TEXT DEFAULT 'Trống',
                note TEXT
            );

            CREATE TABLE IF NOT EXISTS customers(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                phone TEXT,
                id_number TEXT,
                address TEXT,
                note TEXT,
                created_at TEXT
            );

            CREATE TABLE IF NOT EXISTS rentals(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                customer_id INTEGER,
                vehicle_id INTEGER,
                booking_time TEXT,
                start_time TEXT,
                planned_days INTEGER DEFAULT 1,
                due_time TEXT,
                actual_return TEXT,
                price_per_day REAL DEFAULT 0,
                delivery_fee REAL DEFAULT 0,
                deposit REAL DEFAULT 0,
                rental_total REAL DEFAULT 0,
                extra_fee REAL DEFAULT 0,
                damage_fee REAL DEFAULT 0,
                fuel_fee REAL DEFAULT 0,
                discount REAL DEFAULT 0,
                final_total REAL DEFAULT 0,
                refund REAL DEFAULT 0,
                status TEXT DEFAULT 'Đang thuê',
                note TEXT,
                delivery_note TEXT
            );

            CREATE TABLE IF NOT EXISTS extensions(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                rental_id INTEGER,
                old_due TEXT,
                new_due TEXT,
                added_days INTEGER,
                added_amount REAL,
                created_at TEXT
            );

            CREATE TABLE IF NOT EXISTS expenses(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                expense_date TEXT,
                category TEXT,
                amount REAL DEFAULT 0,
                note TEXT
            );

            CREATE TABLE IF NOT EXISTS app_settings(
                key TEXT PRIMARY KEY,
                value TEXT
            );
            """
        )

        self.add_columns()
        self.c.commit()

    def add_columns(self):

        tables = {
            "rentals": {
                "delivery_fee": "REAL DEFAULT 0",
                "delivery_note": "TEXT",
                "refund": "REAL DEFAULT 0",
            },
            "customers": {
                "created_at": "TEXT"
            }
        }

        for table, cols in tables.items():

            exists = {
                r["name"]
                for r in self.c.execute(
                    f"PRAGMA table_info({table})"
                )
            }

            for name, definition in cols.items():

                if name not in exists:

                    self.c.execute(
                        f"""
                        ALTER TABLE {table}
                        ADD COLUMN {name} {definition}
                        """
                    )

        self.c.commit()

    def migrate_money(self):

        marker = self.one(
            "SELECT value FROM app_settings WHERE key='money_unit'"
        )

        if marker:
            return

        self.c.execute(
            """
            UPDATE vehicles
            SET default_price =
                COALESCE(default_price,0) * 1000
            """
        )

        columns = [
            "price_per_day",
            "delivery_fee",
            "deposit",
            "rental_total",
            "extra_fee",
            "damage_fee",
            "fuel_fee",
            "discount",
            "final_total",
            "refund"
        ]

        for col in columns:

            self.c.execute(
                f"""
                UPDATE rentals
                SET {col}=COALESCE({col},0)*1000
                """
            )

        self.c.execute(
            """
            UPDATE extensions
            SET added_amount =
                COALESCE(added_amount,0)*1000
            """
        )

        self.c.execute(
            """
            UPDATE expenses
            SET amount =
                COALESCE(amount,0)*1000
            """
        )

        self.c.execute(
            """
            INSERT OR REPLACE INTO app_settings(key,value)
            VALUES('money_unit','VND')
            """
        )

        self.c.commit()

    def all(self, sql, p=()):
        return self.c.execute(sql, p).fetchall()

    def one(self, sql, p=()):
        return self.c.execute(sql, p).fetchone()

    def run(self, sql, p=()):

        cur = self.c.execute(sql, p)

        self.c.commit()

        return cur.lastrowid


@st.cache_resource
def get_db():
    return DB()


db = get_db()


# =========================================================
# HEADER
# =========================================================

st.markdown(
    f"""
<div class="main-header">
    <h1>🏍️ QUẢN LÝ CHO THUÊ XE MÁY</h1>
    <p>
        Quản lý xe • Giao xe • Trả xe • Gia hạn • Chi phí • Doanh thu
        &nbsp;&nbsp; | &nbsp;&nbsp;
        {now().strftime("%d/%m/%Y %H:%M:%S")}
    </p>
</div>
""",
    unsafe_allow_html=True,
)


# =========================================================
# SIDEBAR
# =========================================================

with st.sidebar:

    st.markdown(
        """
        <h2 style="color:white!important;">
        🏍️ QUẢN LÝ XE
        </h2>
        """,
        unsafe_allow_html=True
    )

    menu = st.radio(
        "CHỨC NĂNG",
        [
            "📊 Tổng quan",
            "🚚 Giao / Trả xe",
            "🏍️ Quản lý xe",
            "💸 Chi phí",
            "📈 Doanh thu / Excel",
        ],
    )

    st.divider()

    st.caption(
        f"Cơ sở dữ liệu:\n{os.path.basename(DB_FILE)}"
    )


# =========================================================
# QUERY
# =========================================================

def active_rentals():

    return db.all(
        """
        SELECT
            r.*,
            c.name customer_name,
            c.phone customer_phone,
            c.id_number customer_id_number,
            c.address customer_address,

            v.code vehicle_code,
            v.plate vehicle_plate,
            v.model vehicle_model,
            v.color vehicle_color

        FROM rentals r

        LEFT JOIN customers c
            ON c.id=r.customer_id

        LEFT JOIN vehicles v
            ON v.id=r.vehicle_id

        WHERE r.status IN ('Đang thuê','Gia hạn')

        ORDER BY r.due_time
        """
    )


def rental_full(rid):

    return db.one(
        """
        SELECT
            r.*,

            c.name customer_name,
            c.phone customer_phone,
            c.id_number customer_id_number,
            c.address customer_address,

            v.code vehicle_code,
            v.plate vehicle_plate,
            v.model vehicle_model,
            v.color vehicle_color

        FROM rentals r

        LEFT JOIN customers c
            ON c.id=r.customer_id

        LEFT JOIN vehicles v
            ON v.id=r.vehicle_id

        WHERE r.id=?
        """,
        (rid,)
    )


# =========================================================
# TỔNG QUAN
# =========================================================

if menu == "📊 Tổng quan":

    active = active_rentals()

    overdue = [
        r for r in active
        if is_overdue(r)
    ]

    free = db.one(
        """
        SELECT COUNT(*) total
        FROM vehicles
        WHERE status='Trống'
        """
    )["total"]

    revenue = db.one(
        """
        SELECT COALESCE(SUM(final_total),0) total
        FROM rentals

        WHERE status='Đã trả'

        AND substr(actual_return,7,4)=?

        AND substr(actual_return,4,2)=?
        """,
        (
            str(now().year),
            f"{now().month:02d}"
        )
    )["total"]

    c1, c2, c3, c4 = st.columns(4)

    cards = [
        (c1, "🚗 XE ĐANG THUÊ", len(active)),
        (c2, "⚠️ ĐƠN QUÁ HẠN", len(overdue)),
        (c3, "🟢 XE ĐANG TRỐNG", free),
        (c4, "💰 DOANH THU THÁNG", money(revenue)),
    ]

    for col, title, value in cards:

        with col:

            st.markdown(
                f"""
                <div class="dashboard-card">
                    <div class="dashboard-title">
                        {title}
                    </div>
                    <div class="dashboard-value">
                        {value}
                    </div>
                </div>
                """,
                unsafe_allow_html=True
            )

    st.write("")

    st.markdown(
        '<div class="section-title">🚚 ĐƠN ĐANG THUÊ</div>',
        unsafe_allow_html=True
    )

    if active:

        data = []

        for r in active:

            late = is_overdue(r)

            data.append(
                {
                    "Đơn": r["id"],
                    "Mã xe": r["vehicle_code"],
                    "Biển số": r["vehicle_plate"],
                    "Khách hàng": r["customer_name"] or "",
                    "Điện thoại": r["customer_phone"] or "",
                    "Ngày giao": r["start_time"],
                    "Hạn trả": r["due_time"],
                    "Tổng": money(r["rental_total"]),
                    "Trạng thái":
                        "⚠ QUÁ HẠN"
                        if late
                        else r["status"]
                }
            )

        st.dataframe(
            data,
            use_container_width=True,
            hide_index=True
        )

    else:

        st.info("Hiện chưa có xe đang thuê.")


# =========================================================
# QUẢN LÝ XE
# =========================================================

elif menu == "🏍️ Quản lý xe":

    st.markdown(
        '<div class="section-title">🏍️ QUẢN LÝ XE</div>',
        unsafe_allow_html=True
    )

    # -------------------------------
    # THÊM XE
    # -------------------------------

    with st.container(border=True):

        st.subheader("➕ Thêm xe")

        c1, c2, c3 = st.columns(3)

        with c1:
            code = st.text_input(
                "Mã xe",
                placeholder="VD: XE001"
            )

            plate = st.text_input(
                "Biển số",
                placeholder="VD: 75A1-12345"
            )

        with c2:

            model = st.text_input(
                "Dòng xe",
                placeholder="Honda Vision"
            )

            color = st.text_input(
                "Màu xe",
                placeholder="Đỏ / Đen / Trắng"
            )

        with c3:

            price = st.text_input(
                "Giá thuê / 24 giờ",
                value="0"
            )

            note = st.text_area(
                "Ghi chú",
                height=65,
                placeholder="Ghi chú ngắn..."
            )

        if st.button(
            "➕ THÊM XE",
            type="primary",
            use_container_width=True
        ):

            try:

                code = code.strip()
                plate = plate.strip()
                model = model.strip()
                color = color.strip()

                if not code:
                    raise ValueError(
                        "Chưa nhập mã xe."
                    )

                if not plate:
                    raise ValueError(
                        "Chưa nhập biển số."
                    )

                if not model:
                    raise ValueError(
                        "Chưa nhập dòng xe."
                    )

                price_value = money_input(price)

                db.run(
                    """
                    INSERT INTO vehicles(
                        code,
                        plate,
                        model,
                        color,
                        default_price,
                        status,
                        note
                    )
                    VALUES(?,?,?,?,?,'Trống',?)
                    """,
                    (
                        code,
                        plate,
                        model,
                        color,
                        price_value,
                        note.strip()
                    )
                )

                st.success(
                    f"Đã thêm xe {code}."
                )

                st.rerun()

            except sqlite3.IntegrityError:

                st.error(
                    "Mã xe hoặc biển số đã tồn tại."
                )

            except Exception as e:

                st.error(str(e))

    st.write("")

    # -------------------------------
    # DANH SÁCH XE
    # -------------------------------

    vehicles = db.all(
        """
        SELECT *
        FROM vehicles
        ORDER BY code
        """
    )

    data = []

    for v in vehicles:

        data.append(
            {
                "ID": v["id"],
                "Mã xe": v["code"],
                "Biển số": v["plate"],
                "Dòng xe": v["model"] or "",
                "Màu": v["color"] or "",
                "Giá / 24h": money(v["default_price"]),
                "Trạng thái": v["status"],
                "Ghi chú": v["note"] or "",
            }
        )

    st.dataframe(
        data,
        use_container_width=True,
        hide_index=True
    )

    st.divider()

    # -------------------------------
    # XÓA XE
    # -------------------------------

    if vehicles:

        codes = [
            v["code"]
            for v in vehicles
        ]

        selected_code = st.selectbox(
            "Chọn xe để xóa",
            codes
        )

        if st.button(
            "🗑 XÓA XE",
            type="secondary"
        ):

            v = db.one(
                """
                SELECT *
                FROM vehicles
                WHERE code=?
                """,
                (selected_code,)
            )

            if v["status"] != "Trống":

                st.error(
                    "Không thể xóa xe đang được thuê."
                )

            else:

                db.run(
                    """
                    DELETE FROM vehicles
                    WHERE id=?
                    """,
                    (v["id"],)
                )

                st.success(
                    f"Đã xóa xe {selected_code}."
                )

                st.rerun()


# =========================================================
# GIAO / TRẢ XE
# =========================================================

elif menu == "🚚 Giao / Trả xe":

    tab_giao, tab_dang_thue = st.tabs(
        [
            "🚚 Giao xe",
            "📋 Xe đang thuê"
        ]
    )

    # =====================================================
    # GIAO XE
    # =====================================================

    with tab_giao:

        st.markdown(
            '<div class="section-title">🚚 TẠO HỢP ĐỒNG + GIAO XE</div>',
            unsafe_allow_html=True
        )

        vehicles = db.all(
            """
            SELECT *
            FROM vehicles
            WHERE status='Trống'
            ORDER BY code
            """
        )

        if not vehicles:

            st.warning(
                "Hiện không có xe trống để giao."
            )

        else:

            vehicle_options = {
                f"{v['code']} | {v['plate']} | {v['model']}": v
                for v in vehicles
            }

            with st.form("rental_form"):

                st.subheader("👤 Thông tin khách hàng")

                c1, c2 = st.columns(2)

                with c1:

                    name = st.text_input(
                        "Tên khách hàng *"
                    )

                    phone = st.text_input(
                        "Điện thoại *"
                    )

                    id_number = st.text_input(
                        "CCCD / Hộ chiếu"
                    )

                with c2:

                    address = st.text_input(
                        "Địa chỉ"
                    )

                    selected_vehicle = st.selectbox(
                        "Xe *",
                        list(vehicle_options.keys())
                    )

                    vehicle = vehicle_options[
                        selected_vehicle
                    ]

                st.divider()

                st.subheader(
                    "🏍️ Thông tin thuê xe"
                )

                c1, c2, c3 = st.columns(3)

                with c1:

                    default_price = int(
                        num(vehicle["default_price"])
                    )

                    price = st.text_input(
                        "Giá thuê / 24 giờ",
                        value=str(default_price)
                    )

                with c2:

                    days = st.number_input(
                        "Số ngày / 24 giờ",
                        min_value=1,
                        max_value=365,
                        value=1,
                        step=1
                    )

                with c3:

                    deposit = st.text_input(
                        "Tiền cọc",
                        value="0"
                    )

                c1, c2 = st.columns(2)

                with c1:

                    start_date = st.date_input(
                        "Ngày giao xe",
                        value=now().date(),
                        format="DD/MM/YYYY"
                    )

                with c2:

                    start_time = st.time_input(
                        "Giờ giao xe",
                        value=now().time().replace(
                            second=0,
                            microsecond=0
                        )
                    )

                note = st.text_area(
                    "Ghi chú",
                    height=70,
                    placeholder="Ghi chú về khách hàng / xe..."
                )

                st.info(
                    "ℹ️ Thời gian thuê được tính theo 24 giờ kể từ thời điểm giao xe."
                )

                submit = st.form_submit_button(
                    "🚚 TẠO HỢP ĐỒNG + GIAO XE",
                    type="primary",
                    use_container_width=True
                )

            if submit:

                try:

                    name = name.strip()
                    phone = phone.strip()

                    if not name:
                        raise ValueError(
                            "Chưa nhập tên khách hàng."
                        )

                    if not phone:
                        raise ValueError(
                            "Chưa nhập số điện thoại."
                        )

                    price_value = money_input(
                        price
                    )

                    deposit_value = money_input(
                        deposit
                    )

                    start_dt = datetime.combine(
                        start_date,
                        start_time
                    )

                    due_dt = (
                        start_dt
                        + timedelta(
                            hours=24 * int(days)
                        )
                    )

                    # -----------------------------
                    # KHÁCH HÀNG
                    # -----------------------------

                    customer = db.one(
                        """
                        SELECT *
                        FROM customers
                        WHERE phone=?
                        ORDER BY id DESC
                        LIMIT 1
                        """,
                        (phone,)
                    )

                    if customer:

                        customer_id = customer["id"]

                        db.run(
                            """
                            UPDATE customers

                            SET
                                name=?,
                                id_number=?,
                                address=?,
                                note=?

                            WHERE id=?
                            """,
                            (
                                name,
                                id_number,
                                address,
                                note,
                                customer_id
                            )
                        )

                    else:

                        customer_id = db.run(
                            """
                            INSERT INTO customers(
                                name,
                                phone,
                                id_number,
                                address,
                                note,
                                created_at
                            )
                            VALUES(?,?,?,?,?,?)
                            """,
                            (
                                name,
                                phone,
                                id_number,
                                address,
                                note,
                                now().strftime(FMT)
                            )
                        )

                    # -----------------------------
                    # TIỀN
                    # -----------------------------

                    rental_total = (
                        price_value
                        * int(days)
                    )

                    # -----------------------------
                    # TẠO ĐƠN
                    # -----------------------------

                    rental_id = db.run(
                        """
                        INSERT INTO rentals(

                            customer_id,
                            vehicle_id,

                            booking_time,
                            start_time,

                            planned_days,
                            due_time,

                            price_per_day,
                            deposit,

                            rental_total,
                            final_total,

                            status,
                            note,
                            delivery_note

                        )
                        VALUES(
                            ?,?,
                            ?,?,
                            ?,?,
                            ?,?,
                            ?,?,
                            'Đang thuê',
                            ?,?
                        )
                        """,
                        (
                            customer_id,
                            vehicle["id"],

                            now().strftime(FMT),
                            start_dt.strftime(FMT),

                            int(days),
                            due_dt.strftime(FMT),

                            price_value,
                            deposit_value,

                            rental_total,
                            rental_total,

                            note,
                            note
                        )
                    )

                    # -----------------------------
                    # CẬP NHẬT XE
                    # -----------------------------

                    db.run(
                        """
                        UPDATE vehicles

                        SET status='Đang thuê'

                        WHERE id=?
                        """,
                        (vehicle["id"],)
                    )

                    st.success(
                        f"🎉 Giao xe thành công — Đơn #{rental_id}"
                    )

                    c1, c2, c3 = st.columns(3)

                    with c1:

                        st.metric(
                            "Xe",
                            vehicle["code"]
                        )

                    with c2:

                        st.metric(
                            "Tiền thuê",
                            money(rental_total)
                        )

                    with c3:

                        st.metric(
                            "Tiền cọc",
                            money(deposit_value)
                        )

                    st.info(
                        f"""
                        **Khách:** {name}

                        **Xe:** {vehicle["code"]} - {vehicle["plate"]}

                        **Giao:** {start_dt.strftime(FMT)}

                        **Hạn trả:** {due_dt.strftime(FMT)}

                        **Đơn:** #{rental_id}
                        """
                    )

                except Exception as e:

                    st.error(
                        f"Không thể giao xe: {e}"
                    )

    # =====================================================
    # XE ĐANG THUÊ
    # =====================================================

    with tab_dang_thue:

        st.markdown(
            '<div class="section-title">📋 XE ĐANG THUÊ</div>',
            unsafe_allow_html=True
        )

        active = active_rentals()

        if not active:

            st.info(
                "Hiện không có xe đang thuê."
            )

        else:

            data = []

            for r in active:

                late = is_overdue(r)

                data.append(
                    {
                        "Đơn": r["id"],
                        "Mã xe": r["vehicle_code"],
                        "Biển số": r["vehicle_plate"],
                        "Khách hàng":
                            r["customer_name"] or "",
                        "Điện thoại":
                            r["customer_phone"] or "",
                        "Ngày giao":
                            r["start_time"],
                        "Hạn trả":
                            r["due_time"],
                        "Giá / 24h":
                            money(r["price_per_day"]),
                        "Cọc":
                            money(r["deposit"]),
                        "Trạng thái":
                            "⚠ QUÁ HẠN"
                            if late
                            else r["status"]
                    }
                )

            st.dataframe(
                data,
                use_container_width=True,
                hide_index=True
            )

            st.divider()

            ids = [
                r["id"]
                for r in active
            ]

            selected_id = st.selectbox(
                "Chọn đơn cần xử lý",
                ids,
                format_func=lambda x:
                    f"Đơn #{x}"
            )

            r = rental_full(
                selected_id
            )

            c1, c2, c3 = st.columns(3)

            with c1:

                if st.button(
                    "⏱ GIA HẠN",
                    use_container_width=True
                ):

                    st.session_state[
                        "extend_id"
                    ] = selected_id

            with c2:

                if st.button(
                    "↩ TRẢ XE",
                    type="primary",
                    use_container_width=True
                ):

                    st.session_state[
                        "return_id"
                    ] = selected_id

            with c3:

                if st.button(
                    "🧾 XUẤT PHIẾU",
                    use_container_width=True
                ):

                    st.session_state[
                        "receipt_id"
                    ] = selected_id

            # =================================================
            # GIA HẠN
            # =================================================

            if (
                "extend_id"
                in st.session_state
                and st.session_state["extend_id"]
                == selected_id
            ):

                st.divider()

                st.subheader(
                    f"⏱ Gia hạn đơn #{selected_id}"
                )

                c1, c2 = st.columns(2)

                with c1:

                    extend_days = st.number_input(
                        "Số ngày gia hạn",
                        min_value=1,
                        max_value=365,
                        value=1
                    )

                with c2:

                    extend_price = st.text_input(
                        "Giá / 24 giờ",
                        value=str(
                            int(
                                num(
                                    r["price_per_day"]
                                )
                            )
                        )
                    )

                c1, c2 = st.columns(2)

                with c1:

                    if st.button(
                        "💾 LƯU GIA HẠN",
                        type="primary",
                        use_container_width=True
                    ):

                        try:

                            p = money_input(
                                extend_price
                            )

                            old_due = dt(
                                r["due_time"]
                            )

                            new_due = (
                                old_due
                                + timedelta(
                                    hours=24
                                    * int(extend_days)
                                )
                            )

                            added_amount = (
                                p
                                * int(extend_days)
                            )

                            db.run(
                                """
                                INSERT INTO extensions(
                                    rental_id,
                                    old_due,
                                    new_due,
                                    added_days,
                                    added_amount,
                                    created_at
                                )
                                VALUES(?,?,?,?,?,?)
                                """,
                                (
                                    selected_id,
                                    r["due_time"],
                                    new_due.strftime(FMT),
                                    int(extend_days),
                                    added_amount,
                                    now().strftime(FMT)
                                )
                            )

                            db.run(
                                """
                                UPDATE rentals

                                SET
                                    planned_days =
                                        planned_days + ?,

                                    due_time=?,

                                    rental_total =
                                        rental_total + ?,

                                    final_total =
                                        final_total + ?,

                                    status='Gia hạn'

                                WHERE id=?
                                """,
                                (
                                    int(extend_days),
                                    new_due.strftime(FMT),
                                    added_amount,
                                    added_amount,
                                    selected_id
                                )
                            )

                            del st.session_state[
                                "extend_id"
                            ]

                            st.success(
                                f"""
                                Đã gia hạn thành công.

                                Hạn mới:
                                {new_due.strftime(FMT)}

                                Tiền gia hạn:
                                {money(added_amount)}
                                """
                            )

                            st.rerun()

                        except Exception as e:

                            st.error(str(e))

                with c2:

                    if st.button(
                        "HỦY",
                        use_container_width=True
                    ):

                        del st.session_state[
                            "extend_id"
                        ]

                        st.rerun()

            # =================================================
            # TRẢ XE
            # =================================================

            if (
                "return_id"
                in st.session_state
                and st.session_state["return_id"]
                == selected_id
            ):

                st.divider()

                st.subheader(
                    f"↩ Trả xe — Đơn #{selected_id}"
                )

                c1, c2 = st.columns(2)

                with c1:

                    return_date = st.date_input(
                        "Ngày trả",
                        value=now().date()
                    )

                with c2:

                    return_time = st.time_input(
                        "Giờ trả",
                        value=now().time().replace(
                            second=0,
                            microsecond=0
                        )
                    )

                c1, c2, c3 = st.columns(3)

                with c1:

                    refund = st.text_input(
                        "Hoàn tiền khách",
                        value="0"
                    )

                with c2:

                    damage = st.text_input(
                        "Phí hư hỏng",
                        value="0"
                    )

                with c3:

                    fuel = st.text_input(
                        "Phí xăng",
                        value="0"
                    )

                discount = st.text_input(
                    "Giảm giá",
                    value="0"
                )

                st.caption(
                    "Hoàn tiền khách được trừ khỏi tổng tiền."
                )

                c1, c2 = st.columns(2)

                with c1:

                    if st.button(
                        "✓ XÁC NHẬN TRẢ XE",
                        type="primary",
                        use_container_width=True
                    ):

                        try:

                            return_dt = datetime.combine(
                                return_date,
                                return_time
                            )

                            start_dt = dt(
                                r["start_time"]
                            )

                            if return_dt < start_dt:

                                raise ValueError(
                                    "Ngày giờ trả không được trước ngày giao."
                                )

                            refund_value = money_input(
                                refund
                            )

                            damage_value = money_input(
                                damage
                            )

                            fuel_value = money_input(
                                fuel
                            )

                            discount_value = money_input(
                                discount
                            )

                            total = (
                                num(
                                    r["rental_total"]
                                )
                                + damage_value
                                + fuel_value
                                - discount_value
                                - refund_value
                            )

                            if total < 0:

                                raise ValueError(
                                    "Tổng tiền không thể nhỏ hơn 0."
                                )

                            db.run(
                                """
                                UPDATE rentals

                                SET
                                    actual_return=?,
                                    refund=?,
                                    damage_fee=?,
                                    fuel_fee=?,
                                    discount=?,
                                    final_total=?,
                                    status='Đã trả'

                                WHERE id=?
                                """,
                                (
                                    return_dt.strftime(FMT),
                                    refund_value,
                                    damage_value,
                                    fuel_value,
                                    discount_value,
                                    total,
                                    selected_id
                                )
                            )

                            db.run(
                                """
                                UPDATE vehicles

                                SET status='Trống'

                                WHERE id=?
                                """,
                                (r["vehicle_id"],)
                            )

                            del st.session_state[
                                "return_id"
                            ]

                            st.success(
                                f"""
                                🎉 Trả xe thành công.

                                Tổng tiền:
                                {money(total)}
                                """
                            )

                            st.rerun()

                        except Exception as e:

                            st.error(
                                f"Lỗi trả xe: {e}"
                            )

                with c2:

                    if st.button(
                        "HỦY",
                        use_container_width=True
                    ):

                        del st.session_state[
                            "return_id"
                        ]

                        st.rerun()

            # =================================================
            # PHIẾU
            # =================================================

            if (
                "receipt_id"
                in st.session_state
                and st.session_state["receipt_id"]
                == selected_id
            ):

                st.divider()

                if not PIL_OK:

                    st.error(
                        "Chưa cài Pillow."
                    )

                else:

                    try:

                        path = create_receipt_image(
                            selected_id
                        )

                        with open(
                            path,
                            "rb"
                        ) as f:

                            st.download_button(
                                "⬇️ TẢI PHIẾU GIAO XE",
                                data=f.read(),
                                file_name=os.path.basename(
                                    path
                                ),
                                mime="image/png",
                                use_container_width=True
                            )

                    except Exception as e:

                        st.error(str(e))


# =========================================================
# CHI PHÍ
# =========================================================

elif menu == "💸 Chi phí":

    st.markdown(
        '<div class="section-title">💸 QUẢN LÝ CHI PHÍ</div>',
        unsafe_allow_html=True
    )

    with st.container(border=True):

        st.subheader("➕ Nhập chi phí")

        c1, c2, c3 = st.columns(3)

        with c1:

            expense_date = st.date_input(
                "Ngày",
                value=now().date()
            )

        with c2:

            category = st.text_input(
                "Loại chi phí",
                placeholder="Xăng, sửa xe..."
            )

        with c3:

            amount = st.text_input(
                "Số tiền",
                value="0"
            )

        note = st.text_input(
            "Ghi chú"
        )

        if st.button(
            "➕ LƯU CHI PHÍ",
            type="primary",
            use_container_width=True
        ):

            try:

                amount_value = money_input(
                    amount
                )

                db.run(
                    """
                    INSERT INTO expenses(
                        expense_date,
                        category,
                        amount,
                        note
                    )
                    VALUES(?,?,?,?)
                    """,
                    (
                        expense_date.strftime(
                            "%d/%m/%Y"
                        ),
                        category.strip()
                        or "Khác",
                        amount_value,
                        note.strip()
                    )
                )

                st.success(
                    "Đã lưu chi phí."
                )

                st.rerun()

            except Exception as e:

                st.error(str(e))

    expenses = db.all(
        """
        SELECT *
        FROM expenses
        ORDER BY id DESC
        """
    )

    data = []

    for e in expenses:

        data.append(
            {
                "Ngày":
                    e["expense_date"],
                "Loại":
                    e["category"],
                "Số tiền":
                    money(e["amount"]),
                "Ghi chú":
                    e["note"] or "",
            }
        )

    st.dataframe(
        data,
        use_container_width=True,
        hide_index=True
    )


# =========================================================
# DOANH THU
# =========================================================

elif menu == "📈 Doanh thu / Excel":

    st.markdown(
        '<div class="section-title">📈 DOANH THU / EXCEL</div>',
        unsafe_allow_html=True
    )

    c1, c2 = st.columns(2)

    with c1:

        month = st.selectbox(
            "Tháng",
            range(1, 13),
            index=now().month - 1,
            format_func=lambda x:
                f"Tháng {x:02d}"
        )

    with c2:

        year = st.number_input(
            "Năm",
            min_value=2000,
            max_value=2100,
            value=now().year
        )

    rows = db.all(
        """
        SELECT
            r.*,

            c.name customer_name,
            c.phone,

            v.code,
            v.plate,
            v.model

        FROM rentals r

        LEFT JOIN customers c
            ON c.id=r.customer_id

        LEFT JOIN vehicles v
            ON v.id=r.vehicle_id

        WHERE
            r.status='Đã trả'

        AND substr(
            r.actual_return,
            7,
            4
        )=?

        AND substr(
            r.actual_return,
            4,
            2
        )=?

        ORDER BY r.actual_return
        """,
        (
            str(year),
            f"{month:02d}"
        )
    )

    expenses = db.all(
        """
        SELECT *
        FROM expenses

        WHERE
            substr(
                expense_date,
                7,
                4
            )=?

        AND substr(
            expense_date,
            4,
            2
        )=?

        ORDER BY expense_date
        """,
        (
            str(year),
            f"{month:02d}"
        )
    )

    revenue = sum(
        num(r["final_total"])
        for r in rows
    )

    cost = sum(
        num(e["amount"])
        for e in expenses
    )

    profit = revenue - cost

    c1, c2, c3, c4 = st.columns(4)

    with c1:
        st.metric(
            "Số đơn đã trả",
            len(rows)
        )

    with c2:
        st.metric(
            "Doanh thu",
            money(revenue)
        )

    with c3:
        st.metric(
            "Chi phí",
            money(cost)
        )

    with c4:
        st.metric(
            "Lợi nhuận",
            money(profit)
        )

    st.divider()

    st.subheader(
        f"📋 Chi tiết đơn — {month:02d}/{year}"
    )

    data = []

    for r in rows:

        data.append(
            {
                "Đơn": r["id"],
                "Ngày trả": r["actual_return"],
                "Mã xe": r["code"],
                "Biển số": r["plate"],
                "Dòng xe": r["model"],
                "Khách hàng":
                    r["customer_name"] or "",
                "Điện thoại":
                    r["phone"] or "",
                "Ngày giao":
                    r["start_time"],
                "Hạn trả":
                    r["due_time"],
                "Giá / 24h":
                    money(r["price_per_day"]),
                "Cọc":
                    money(r["deposit"]),
                "Hoàn tiền":
                    money(r["refund"]),
                "Phí hư hỏng":
                    money(r["damage_fee"]),
                "Phí xăng":
                    money(r["fuel_fee"]),
                "Giảm giá":
                    money(r["discount"]),
                "Tổng tiền":
                    money(r["final_total"]),
            }
        )

    st.dataframe(
        data,
        use_container_width=True,
        hide_index=True
    )

    st.subheader("💸 Chi phí")

    expense_data = []

    for e in expenses:

        expense_data.append(
            {
                "Ngày":
                    e["expense_date"],
                "Loại":
                    e["category"],
                "Số tiền":
                    money(e["amount"]),
                "Ghi chú":
                    e["note"] or "",
            }
        )

    st.dataframe(
        expense_data,
        use_container_width=True,
        hide_index=True
    )

    # =====================================================
    # EXCEL
    # =====================================================

    if not XLSX_OK:

        st.warning(
            "Chưa cài openpyxl nên chưa thể xuất Excel."
        )

    else:

        if st.button(
            "📗 TẠO FILE EXCEL",
            type="primary",
            use_container_width=True
        ):

            try:

                excel_data = create_excel(
                    rows,
                    expenses,
                    month,
                    year
                )

                st.download_button(
                    "⬇️ TẢI BÁO CÁO EXCEL",
                    data=excel_data,
                    file_name=(
                        f"Doanh_thu_"
                        f"{year}_"
                        f"{month:02d}.xlsx"
                    ),
                    mime=(
                        "application/"
                        "vnd.openxmlformats-officedocument"
                        ".spreadsheetml.sheet"
                    ),
                    use_container_width=True
                )

            except Exception as e:

                st.error(
                    f"Lỗi xuất Excel: {e}"
                )


# =========================================================
# TẠO PHIẾU GIAO XE
# =========================================================

def create_receipt_image(rid):

    r = rental_full(rid)

    if not r:

        raise ValueError(
            "Không tìm thấy đơn."
        )

    if not PIL_OK:

        raise ValueError(
            "Máy chưa cài Pillow."
        )

    start_dt = dt(
        r["start_time"]
    )

    date_str = start_dt.strftime(
        "%d-%m-%Y"
    )

    png = os.path.join(
        PHIEU_DIR,
        f"PHIEU GIAO XE SỐ "
        f"{rid}_{date_str}.png"
    )

    width = 1100
    height = 1850

    im = Image.new(
        "RGB",
        (width, height),
        "white"
    )

    d = ImageDraw.Draw(im)

    normal_fonts = [
        "C:/Windows/Fonts/arial.ttf",
        "C:/Windows/Fonts/segoeui.ttf",
        "C:/Windows/Fonts/tahoma.ttf"
    ]

    bold_fonts = [
        "C:/Windows/Fonts/arialbd.ttf",
        "C:/Windows/Fonts/segoeuib.ttf",
        "C:/Windows/Fonts/tahomabd.ttf"
    ]

    nf = next(
        (
            x for x in normal_fonts
            if os.path.exists(x)
        ),
        None
    )

    bf = next(
        (
            x for x in bold_fonts
            if os.path.exists(x)
        ),
        nf
    )

    if nf:

        title = ImageFont.truetype(
            bf,
            34
        )

        head = ImageFont.truetype(
            bf,
            23
        )

        font = ImageFont.truetype(
            nf,
            20
        )

        small = ImageFont.truetype(
            nf,
            17
        )

    else:

        title = head = font = small = (
            ImageFont.load_default()
        )

    x = 60
    y = 35

    d.text(
        (x, y),
        "PHIẾU GIAO XE MÁY",
        font=title,
        fill="#172033"
    )

    y += 60

    d.text(
        (x, y),
        f"PHIẾU SỐ {rid} - {date_str}",
        font=head,
        fill="#334155"
    )

    y += 45

    d.line(
        (x, y, width - x, y),
        fill="#94a3b8",
        width=2
    )

    y += 30

    sections = [

        (
            "KHÁCH HÀNG",
            [
                (
                    "Tên",
                    r["customer_name"] or ""
                ),
                (
                    "Điện thoại",
                    r["customer_phone"] or ""
                ),
                (
                    "CCCD/Hộ chiếu",
                    r["customer_id_number"] or ""
                ),
                (
                    "Địa chỉ",
                    r["customer_address"] or ""
                ),
            ]
        ),

        (
            "XE",
            [
                (
                    "Mã xe",
                    r["vehicle_code"]
                ),
                (
                    "Biển số",
                    r["vehicle_plate"]
                ),
                (
                    "Dòng xe",
                    r["vehicle_model"]
                ),
                (
                    "Màu",
                    r["vehicle_color"] or ""
                ),
            ]
        ),

        (
            "THỜI GIAN",
            [
                (
                    "Ngày giờ giao",
                    r["start_time"]
                ),
                (
                    "Ngày giờ hết hạn",
                    r["due_time"]
                ),
                (
                    "Số ngày",
                    f"{r['planned_days']} x 24 giờ"
                ),
            ]
        )
    ]

    for title_text, lines in sections:

        d.text(
            (x, y),
            title_text,
            font=head,
            fill="#1d4ed8"
        )

        y += 38

        for label, value in lines:

            d.text(
                (x + 15, y),
                label,
                font=small,
                fill="#64748b"
            )

            d.text(
                (x + 270, y),
                str(value),
                font=font,
                fill="#111827"
            )

            y += 38

        y += 10

    d.line(
        (x, y, width - x, y),
        fill="#cbd5e1",
        width=2
    )

    y += 25

    d.text(
        (x, y),
        "THANH TOÁN",
        font=head,
        fill="#15803d"
    )

    y += 42

    payment_lines = [
        (
            "Giá / 24 giờ",
            money(r["price_per_day"])
        ),
        (
            "Tiền cọc",
            money(r["deposit"])
        ),
        (
            "TỔNG TIỀN",
            money(r["rental_total"])
        )
    ]

    for label, value in payment_lines:

        d.text(
            (x + 15, y),
            label,
            font=font,
            fill="#111827"
        )

        d.text(
            (x + 650, y),
            value,
            font=font,
            fill="#111827"
        )

        y += 40

    y += 15

    d.line(
        (x, y, width - x, y),
        fill="#cbd5e1",
        width=2
    )

    y += 25

    d.text(
        (x, y),
        "GHI CHÚ",
        font=head,
        fill="#334155"
    )

    y += 38

    note = r["delivery_note"] or ""

    if note:

        for i in range(
            0,
            len(note),
            75
        ):

            d.text(
                (x + 15, y),
                note[i:i + 75],
                font=font,
                fill="#111827"
            )

            y += 30

    else:

        d.text(
            (x + 15, y),
            "Không có",
            font=font,
            fill="#64748b"
        )

        y += 30

    y += 20

    d.text(
        (x, y),
        "LƯU Ý",
        font=head,
        fill="#b91c1c"
    )

    y += 38

    notes = [
        "• Thời gian thuê tính theo 24 giờ kể từ lúc giao xe.",
        "• Đơn quá hạn sẽ tự động hiển thị màu đỏ."
    ]

    for text in notes:

        d.text(
            (x + 15, y),
            text,
            font=small,
            fill="#334155"
        )

        y += 30

    y += 35

    d.text(
        (x + 20, y),
        "Nhân viên: ______________________________",
        font=font,
        fill="#111827"
    )

    d.text(
        (x + 570, y),
        "Khách hàng: __________________",
        font=font,
        fill="#111827"
    )

    y += 60

    d.line(
        (x, y, width - x, y),
        fill="#94a3b8",
        width=2
    )

    y += 25

    # QR

    if os.path.exists(QR_FILE):

        try:

            qr = Image.open(
                QR_FILE
            ).convert("RGB")

            qr_size = 280

            qr.thumbnail(
                (
                    qr_size,
                    qr_size
                ),
                Image.Resampling.LANCZOS
            )

            qr_x = (
                width - qr.width
            ) // 2

            qr_y = y + 30

            text = "MÃ QR THANH TOÁN"

            bbox = d.textbbox(
                (0, 0),
                text,
                font=head
            )

            d.text(
                (
                    (
                        width
                        - (
                            bbox[2]
                            - bbox[0]
                        )
                    ) // 2,
                    y
                ),
                text,
                font=head,
                fill="#1d4ed8"
            )

            im.paste(
                qr,
                (
                    qr_x,
                    qr_y
                )
            )

            text2 = "Quét mã QR để thanh toán"

            bbox2 = d.textbbox(
                (0, 0),
                text2,
                font=small
            )

            d.text(
                (
                    (
                        width
                        - (
                            bbox2[2]
                            - bbox2[0]
                        )
                    ) // 2,
                    qr_y + qr.height + 15
                ),
                text2,
                font=small,
                fill="#475569"
            )

        except Exception:

            d.text(
                (x, y),
                "QR_THANH_TOAN.png không đọc được.",
                font=font,
                fill="#b91c1c"
            )

    else:

        text = "CHƯA CÓ QR THANH TOÁN"

        bbox = d.textbbox(
            (0, 0),
            text,
            font=head
        )

        d.text(
            (
                (
                    width
                    - (
                        bbox[2]
                        - bbox[0]
                    )
                ) // 2,
                y + 20
            ),
            text,
            font=head,
            fill="#b91c1c"
        )

        text2 = (
            "Đặt file QR_THANH_TOAN.png "
            "cạnh file chương trình."
        )

        bbox2 = d.textbbox(
            (0, 0),
            text2,
            font=small
        )

        d.text(
            (
                (
                    width
                    - (
                        bbox2[2]
                        - bbox2[0]
                    )
                ) // 2,
                y + 65
            ),
            text2,
            font=small,
            fill="#64748b"
        )

    im.save(png)

    return png


# =========================================================
# TẠO EXCEL
# =========================================================

def create_excel(
    rows,
    expenses,
    month,
    year
):

    from io import BytesIO

    wb = Workbook()

    ws = wb.active
    ws.title = "Doanh thu"

    headers = [
        "Đơn",
        "Ngày trả",
        "Mã xe",
        "Biển số",
        "Dòng xe",
        "Khách hàng",
        "Điện thoại",
        "Ngày giờ giao",
        "Hạn trả",
        "Giá / 24h (VNĐ)",
        "Cọc (VNĐ)",
        "Hoàn tiền khách (VNĐ)",
        "Phí hư hỏng (VNĐ)",
        "Xăng (VNĐ)",
        "Giảm giá (VNĐ)",
        "Tổng tiền (VNĐ)",
        "Trạng thái"
    ]

    ws.append(headers)

    for c in ws[1]:

        c.font = Font(
            bold=True,
            color="FFFFFF"
        )

        c.fill = PatternFill(
            "solid",
            fgColor="1F4E78"
        )

        c.alignment = Alignment(
            horizontal="center"
        )

    for r in rows:

        ws.append(
            [
                r["id"],
                r["actual_return"],
                r["code"],
                r["plate"],
                r["model"],
                r["customer_name"],
                r["phone"],
                r["start_time"],
                r["due_time"],
                r["price_per_day"],
                r["deposit"],
                r["refund"],
                r["damage_fee"],
                r["fuel_fee"],
                r["discount"],
                r["final_total"],
                r["status"]
            ]
        )

    for row in ws.iter_rows(
        min_row=2
    ):

        for cell in row:

            if 10 <= cell.column <= 16:

                cell.number_format = (
                    '#,##0" đ"'
                )

    for col in ws.columns:

        length = max(
            len(
                str(
                    c.value or ""
                )
            )
            for c in col
        )

        ws.column_dimensions[
            get_column_letter(
                col[0].column
            )
        ].width = min(
            max(length + 2, 12),
            30
        )

    ws.freeze_panes = "A2"

    # -----------------------------------------------------
    # CHI PHÍ
    # -----------------------------------------------------

    ex = wb.create_sheet(
        "Chi phí"
    )

    ex.append(
        [
            "Ngày",
            "Loại",
            "Số tiền (VNĐ)",
            "Ghi chú"
        ]
    )

    for c in ex[1]:

        c.font = Font(
            bold=True,
            color="FFFFFF"
        )

        c.fill = PatternFill(
            "solid",
            fgColor="7F6000"
        )

    for e in expenses:

        ex.append(
            [
                e["expense_date"],
                e["category"],
                e["amount"],
                e["note"] or ""
            ]
        )

    for row in ex.iter_rows(
        min_row=2
    ):

        row[2].number_format = (
            '#,##0" đ"'
        )

    # -----------------------------------------------------
    # TỔNG HỢP
    # -----------------------------------------------------

    sm = wb.create_sheet(
        "Tổng hợp"
    )

    revenue = sum(
        num(r["final_total"])
        for r in rows
    )

    cost = sum(
        num(e["amount"])
        for e in expenses
    )

    profit = revenue - cost

    sm.append(
        [
            "BÁO CÁO DOANH THU",
            f"{month:02d}/{year}"
        ]
    )

    sm.append(
        [
            "Số đơn đã trả",
            len(rows)
        ]
    )

    sm.append(
        [
            "Doanh thu (VNĐ)",
            revenue
        ]
    )

    sm.append(
        [
            "Chi phí (VNĐ)",
            cost
        ]
    )

    sm.append(
        [
            "Lợi nhuận (VNĐ)",
            profit
        ]
    )

    sm["A1"].font = Font(
        bold=True,
        size=15
    )

    for cell in [
        "B3",
        "B4",
        "B5"
    ]:

        sm[cell].number_format = (
            '#,##0" đ"'
        )

    sm.column_dimensions["A"].width = 25
    sm.column_dimensions["B"].width = 25

    output = BytesIO()

    wb.save(output)

    output.seek(0)

    return output.getvalue()