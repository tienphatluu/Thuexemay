# -*- coding: utf-8 -*-
"""
QUẢN LÝ CHO THUÊ XE MÁY V4
- PySide6
- SQLite
- Pillow
- OpenPyXL
- KHÔNG sử dụng Tkinter
"""

import os
import sys
import sqlite3
from datetime import datetime, timedelta

from PySide6.QtCore import Qt, QDate, QTime, QTimer
from PySide6.QtGui import (
    QColor,
    QFont,
    QPixmap,
)
from PySide6.QtWidgets import (
    QApplication,
    QMainWindow,
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QGridLayout,
    QFormLayout,
    QLabel,
    QPushButton,
    QLineEdit,
    QTextEdit,
    QComboBox,
    QSpinBox,
    QDateEdit,
    QTimeEdit,
    QTableWidget,
    QTableWidgetItem,
    QTabWidget,
    QGroupBox,
    QMessageBox,
    QFileDialog,
    QDialog,
    QDialogButtonBox,
    QDoubleSpinBox,
    QHeaderView,
    QAbstractItemView,
)

# ============================================================
# OPTIONAL LIBRARIES
# ============================================================

try:
    from PIL import Image, ImageDraw, ImageFont
    PIL_OK = True
except ImportError:
    PIL_OK = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font as XLFont, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    XLSX_OK = True
except ImportError:
    XLSX_OK = False


# ============================================================
# PATH
# ============================================================

BASE = os.path.dirname(os.path.abspath(__file__))
DB_FILE = os.path.join(BASE, "rental_v2.db")
PHIEU_DIR = os.path.join(BASE, "PHIEU_GIAO_XE")
QR_FILE = os.path.join(BASE, "QR_THANH_TOAN.png")

FMT = "%d/%m/%Y %H:%M"

os.makedirs(PHIEU_DIR, exist_ok=True)


# ============================================================
# UTILITIES
# ============================================================

def now():
    return datetime.now()


def money(value):
    try:
        return f"{float(value):,.0f}".replace(",", ".") + " đ"
    except Exception:
        return "0 đ"


def num(value):
    try:
        return float(value)
    except Exception:
        return 0.0


def parse_money(value):
    if isinstance(value, (int, float)):
        return float(value)

    s = str(value).strip()
    s = s.replace("VNĐ", "")
    s = s.replace("VND", "")
    s = s.replace("đ", "")
    s = s.replace("Đ", "")
    s = s.replace(" ", "")
    s = s.replace(".", "")
    s = s.replace(",", "")

    if not s:
        return 0.0

    try:
        value = float(s)
    except ValueError:
        raise ValueError("Số tiền không hợp lệ.")

    if value < 0:
        raise ValueError("Số tiền không được âm.")

    return value


def parse_datetime(value):
    return datetime.strptime(value.strip(), FMT)


def dt_to_str(value):
    return value.strftime(FMT)


def qdate_to_datetime(date_obj, time_obj):
    return datetime(
        date_obj.year(),
        date_obj.month(),
        date_obj.day(),
        time_obj.hour(),
        time_obj.minute(),
    )


# ============================================================
# DATABASE
# ============================================================

class Database:

    def __init__(self):
        self.conn = sqlite3.connect(DB_FILE)
        self.conn.row_factory = sqlite3.Row

        self.create_tables()
        self.migrate()

    def execute(self, sql, params=()):
        cur = self.conn.execute(sql, params)
        self.conn.commit()
        return cur

    def fetchone(self, sql, params=()):
        return self.conn.execute(sql, params).fetchone()

    def fetchall(self, sql, params=()):
        return self.conn.execute(sql, params).fetchall()

    def create_tables(self):

        self.conn.executescript("""
        CREATE TABLE IF NOT EXISTS vehicles(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT UNIQUE NOT NULL,
            plate TEXT UNIQUE NOT NULL,
            model TEXT,
            color TEXT,
            default_price REAL DEFAULT 0,
            status TEXT DEFAULT 'Trống',
            note TEXT
        );

        CREATE TABLE IF NOT EXISTS customers(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            phone TEXT,
            id_number TEXT,
            address TEXT,
            note TEXT,
            created_at TEXT
        );

        CREATE TABLE IF NOT EXISTS rentals(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            customer_id INTEGER,
            vehicle_id INTEGER,
            booking_time TEXT,
            start_time TEXT,
            planned_days INTEGER DEFAULT 1,
            due_time TEXT,
            actual_return TEXT,

            price_per_day REAL DEFAULT 0,
            delivery_fee REAL DEFAULT 0,
            deposit REAL DEFAULT 0,

            rental_total REAL DEFAULT 0,
            extra_fee REAL DEFAULT 0,
            damage_fee REAL DEFAULT 0,
            fuel_fee REAL DEFAULT 0,
            discount REAL DEFAULT 0,
            final_total REAL DEFAULT 0,
            refund REAL DEFAULT 0,

            status TEXT DEFAULT 'Đang thuê',
            note TEXT,
            delivery_note TEXT
        );

        CREATE TABLE IF NOT EXISTS extensions(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            rental_id INTEGER,
            old_due TEXT,
            new_due TEXT,
            added_days INTEGER,
            added_amount REAL,
            created_at TEXT
        );

        CREATE TABLE IF NOT EXISTS expenses(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            expense_date TEXT,
            category TEXT,
            amount REAL DEFAULT 0,
            note TEXT
        );

        CREATE TABLE IF NOT EXISTS app_settings(
            key TEXT PRIMARY KEY,
            value TEXT
        );
        """)

        self.add_missing_columns()
        self.conn.commit()

    def add_missing_columns(self):

        tables = {
            "rentals": {
                "delivery_fee": "REAL DEFAULT 0",
                "delivery_note": "TEXT",
                "refund": "REAL DEFAULT 0",
            },
            "customers": {
                "created_at": "TEXT",
            },
        }

        for table, columns in tables.items():

            existing = {
                row["name"]
                for row in self.conn.execute(
                    f"PRAGMA table_info({table})"
                )
            }

            for name, definition in columns.items():

                if name not in existing:
                    self.conn.execute(
                        f"ALTER TABLE {table} ADD COLUMN {name} {definition}"
                    )

    def migrate(self):

        marker = self.fetchone(
            "SELECT value FROM app_settings WHERE key='money_unit'"
        )

        if marker:
            return

        # Database cũ của chương trình ban đầu dùng đơn vị nghìn đồng.
        self.conn.execute("""
            UPDATE vehicles
            SET default_price = COALESCE(default_price,0) * 1000
        """)

        money_columns = [
            "price_per_day",
            "delivery_fee",
            "deposit",
            "rental_total",
            "extra_fee",
            "damage_fee",
            "fuel_fee",
            "discount",
            "final_total",
            "refund",
        ]

        for col in money_columns:
            self.conn.execute(
                f"""
                UPDATE rentals
                SET {col}=COALESCE({col},0)*1000
                """
            )

        self.conn.execute("""
            UPDATE extensions
            SET added_amount=COALESCE(added_amount,0)*1000
        """)

        self.conn.execute("""
            UPDATE expenses
            SET amount=COALESCE(amount,0)*1000
        """)

        self.conn.execute("""
            INSERT OR REPLACE INTO app_settings(key,value)
            VALUES('money_unit','VND')
        """)

        self.conn.commit()

    def close(self):
        self.conn.close()


# ============================================================
# STYLE
# ============================================================

STYLE = """
QWidget {
    background-color: #eef2f7;
    color: #0f172a;
    font-family: "Segoe UI";
    font-size: 13px;
}

QMainWindow {
    background-color: #eef2f7;
}

QLabel {
    color: #0f172a;
}

QGroupBox {
    background-color: #f8fafc;
    color: #0f172a;
    border: 1px solid #cbd5e1;
    border-radius: 7px;
    margin-top: 12px;
    padding: 12px;
    font-weight: bold;
}

QGroupBox::title {
    color: #1e293b;
    subcontrol-origin: margin;
    left: 12px;
    padding: 0 5px;
    background-color: #f8fafc;
}

QLineEdit,
QTextEdit,
QComboBox,
QSpinBox,
QDoubleSpinBox,
QDateEdit,
QTimeEdit {
    background-color: #ffffff;
    color: #111827;
    border: 1px solid #94a3b8;
    border-radius: 5px;
    padding: 6px;
    selection-background-color: #2563eb;
    selection-color: white;
}

QLineEdit:focus,
QTextEdit:focus,
QComboBox:focus,
QSpinBox:focus,
QDoubleSpinBox:focus,
QDateEdit:focus,
QTimeEdit:focus {
    border: 2px solid #2563eb;
}

QPushButton {
    background-color: #2563eb;
    color: white;
    border: none;
    border-radius: 5px;
    padding: 8px 14px;
    font-weight: bold;
}

QPushButton:hover {
    background-color: #1d4ed8;
}

QPushButton:pressed {
    background-color: #1e40af;
}

QPushButton[success="true"] {
    background-color: #16a34a;
}

QPushButton[success="true"]:hover {
    background-color: #15803d;
}

QPushButton[danger="true"] {
    background-color: #dc2626;
}

QPushButton[danger="true"]:hover {
    background-color: #b91c1c;
}

QPushButton[warning="true"] {
    background-color: #d97706;
}

QTableWidget {
    background-color: white;
    alternate-background-color: #f8fafc;
    color: #111827;
    gridline-color: #cbd5e1;
    border: 1px solid #cbd5e1;
    selection-background-color: #bfdbfe;
    selection-color: #0f172a;
}

QHeaderView::section {
    background-color: #1e3a5f;
    color: white;
    padding: 8px;
    border: 1px solid #475569;
    font-weight: bold;
}

QTabWidget::pane {
    border: 1px solid #cbd5e1;
    background-color: #eef2f7;
}

QTabBar::tab {
    background-color: #dbe4ef;
    color: #334155;
    padding: 10px 18px;
    margin-right: 2px;
    font-weight: bold;
}

QTabBar::tab:selected {
    background-color: #2563eb;
    color: white;
}

QTabBar::tab:hover {
    background-color: #93c5fd;
    color: #0f172a;
}

QTextEdit {
    background-color: white;
    color: #111827;
}

QScrollBar:vertical {
    background: #e2e8f0;
    width: 12px;
}

QScrollBar::handle:vertical {
    background: #94a3b8;
    border-radius: 5px;
}

QMessageBox {
    background-color: white;
}

QMessageBox QLabel {
    color: #111827;
}
"""


# ============================================================
# MAIN WINDOW
# ============================================================

class MainWindow(QMainWindow):

    def __init__(self):

        super().__init__()

        self.db = Database()

        self.selected_rental = None
        self.selected_vehicle_id = None

        self.setWindowTitle("QUẢN LÝ CHO THUÊ XE MÁY V4")
        self.resize(1450, 900)
        self.setMinimumSize(1150, 700)

        self.setStyleSheet(STYLE)

        self.build_ui()
        self.refresh_all()

        self.timer = QTimer(self)
        self.timer.timeout.connect(self.auto_refresh)
        self.timer.start(30000)

    # ========================================================
    # BASIC HELPERS
    # ========================================================

    def label(self, text):
        lbl = QLabel(text)
        lbl.setStyleSheet("color:#0f172a;")
        return lbl

    def button(self, text, slot, kind=None):

        btn = QPushButton(text)
        btn.clicked.connect(slot)

        if kind == "success":
            btn.setProperty("success", True)

        elif kind == "danger":
            btn.setProperty("danger", True)

        elif kind == "warning":
            btn.setProperty("warning", True)

        return btn

    def show_error(self, text):
        QMessageBox.critical(self, "Lỗi", str(text))

    def show_warning(self, text):
        QMessageBox.warning(self, "Thông báo", str(text))

    def show_info(self, title, text):
        QMessageBox.information(self, title, text)

    # ========================================================
    # UI
    # ========================================================

    def build_ui(self):

        central = QWidget()
        self.setCentralWidget(central)

        root = QVBoxLayout(central)
        root.setContentsMargins(0, 0, 0, 0)

        # HEADER
        header = QWidget()
        header.setFixedHeight(65)
        header.setStyleSheet("""
            QWidget {
                background-color: #172033;
            }
            QLabel {
                color: white;
            }
        """)

        hbox = QHBoxLayout(header)

        title = QLabel("🏍  QUẢN LÝ CHO THUÊ XE MÁY V4")
        title.setFont(QFont("Segoe UI", 18, QFont.Weight.Bold))

        self.clock_label = QLabel()
        self.clock_label.setFont(QFont("Segoe UI", 11))

        hbox.addWidget(title)
        hbox.addStretch()
        hbox.addWidget(self.clock_label)

        root.addWidget(header)

        self.tabs = QTabWidget()
        root.addWidget(self.tabs)

        self.build_dashboard_tab()
        self.build_rental_tab()
        self.build_vehicle_tab()
        self.build_expense_tab()
        self.build_report_tab()

        self.clock_timer = QTimer(self)
        self.clock_timer.timeout.connect(self.update_clock)
        self.clock_timer.start(1000)

        self.update_clock()

    # ========================================================
    # CLOCK
    # ========================================================

    def update_clock(self):

        self.clock_label.setText(
            now().strftime("%d/%m/%Y  %H:%M:%S")
        )

    # ========================================================
    # DASHBOARD
    # ========================================================

    def build_dashboard_tab(self):

        tab = QWidget()
        layout = QVBoxLayout(tab)

        cards = QHBoxLayout()

        self.card_active = self.create_card(
            cards,
            "XE ĐANG THUÊ",
            "0",
            "#2563eb"
        )

        self.card_overdue = self.create_card(
            cards,
            "ĐƠN QUÁ HẠN",
            "0",
            "#dc2626"
        )

        self.card_free = self.create_card(
            cards,
            "XE ĐANG TRỐNG",
            "0",
            "#16a34a"
        )

        self.card_revenue = self.create_card(
            cards,
            "DOANH THU THÁNG",
            "0 đ",
            "#7c3aed"
        )

        layout.addLayout(cards)

        group = QGroupBox("ĐƠN ĐANG THUÊ")

        g = QVBoxLayout(group)

        columns = [
            "Mã xe",
            "Biển số",
            "Khách hàng",
            "Điện thoại",
            "Ngày giờ giao",
            "Hạn trả",
            "Tổng",
            "Trạng thái",
        ]

        self.dashboard_table = QTableWidget(0, len(columns))
        self.dashboard_table.setHorizontalHeaderLabels(columns)
        self.prepare_table(self.dashboard_table)

        g.addWidget(self.dashboard_table)
        layout.addWidget(group)

        self.tabs.addTab(tab, "TỔNG QUAN")

    def create_card(self, layout, title, value, color):

        box = QGroupBox()

        box.setStyleSheet(f"""
            QGroupBox {{
                background-color: white;
                border: 1px solid #cbd5e1;
                border-top: 5px solid {color};
                border-radius: 8px;
                margin-top: 0;
            }}
        """)

        v = QVBoxLayout(box)

        title_lbl = QLabel(title)
        title_lbl.setStyleSheet("color:#64748b;font-weight:bold;")

        value_lbl = QLabel(value)
        value_lbl.setStyleSheet(
            f"color:{color};font-size:22px;font-weight:bold;"
        )

        v.addWidget(title_lbl)
        v.addWidget(value_lbl)

        layout.addWidget(box)

        return value_lbl

    # ========================================================
    # TABLE
    # ========================================================

    def prepare_table(self, table):

        table.setAlternatingRowColors(True)
        table.setSelectionBehavior(
            QAbstractItemView.SelectionBehavior.SelectRows
        )
        table.setSelectionMode(
            QAbstractItemView.SelectionMode.SingleSelection
        )
        table.setEditTriggers(
            QAbstractItemView.EditTrigger.NoEditTriggers
        )
        table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.ResizeToContents
        )
        table.horizontalHeader().setStretchLastSection(True)

    def set_item(self, table, row, col, value, color=None):

        item = QTableWidgetItem(str(value))
        item.setForeground(QColor(color or "#111827"))

        table.setItem(row, col, item)

    # ========================================================
    # RENTAL TAB
    # ========================================================

    def build_rental_tab(self):

        tab = QWidget()
        root = QHBoxLayout(tab)

        # LEFT
        left_group = QGroupBox("TẠO HỢP ĐỒNG + GIAO XE")
        left_group.setFixedWidth(420)

        form = QGridLayout(left_group)

        row = 0

        self.r_name = QLineEdit()
        self.r_phone = QLineEdit()
        self.r_id = QLineEdit()
        self.r_address = QLineEdit()

        for text, widget in [
            ("Tên khách hàng", self.r_name),
            ("Điện thoại", self.r_phone),
            ("CCCD / Hộ chiếu", self.r_id),
            ("Địa chỉ", self.r_address),
        ]:
            form.addWidget(self.label(text), row, 0)
            form.addWidget(widget, row, 1)
            row += 1

        form.addWidget(self.label("Xe"), row, 0)

        self.vehicle_combo = QComboBox()
        self.vehicle_combo.currentIndexChanged.connect(
            self.vehicle_selected
        )

        form.addWidget(self.vehicle_combo, row, 1)
        row += 1

        form.addWidget(self.label("Giá thuê / 24 giờ"), row, 0)

        self.r_price = QLineEdit()
        self.r_price.setText("0")
        form.addWidget(self.r_price, row, 1)
        row += 1

        form.addWidget(self.label("Số ngày / 24 giờ"), row, 0)

        self.r_days = QSpinBox()
        self.r_days.setRange(1, 365)
        self.r_days.setValue(1)

        form.addWidget(self.r_days, row, 1)
        row += 1

        form.addWidget(self.label("Tiền cọc"), row, 0)

        self.r_deposit = QLineEdit()
        self.r_deposit.setText("0")

        form.addWidget(self.r_deposit, row, 1)
        row += 1

        form.addWidget(self.label("Ngày giao xe"), row, 0)

        self.r_date = QDateEdit()
        self.r_date.setCalendarPopup(True)
        self.r_date.setDate(QDate.currentDate())

        form.addWidget(self.r_date, row, 1)
        row += 1

        form.addWidget(self.label("Giờ giao xe"), row, 0)

        self.r_time = QTimeEdit()
        self.r_time.setDisplayFormat("HH:mm")
        self.r_time.setTime(QTime.currentTime())

        form.addWidget(self.r_time, row, 1)
        row += 1

        form.addWidget(self.label("Ghi chú"), row, 0)

        self.r_note = QTextEdit()
        self.r_note.setFixedHeight(90)

        form.addWidget(self.r_note, row, 1)
        row += 1

        create_btn = self.button(
            "🚚 TẠO HỢP ĐỒNG + GIAO XE",
            self.create_rental,
            "success"
        )

        form.addWidget(create_btn, row, 0, 1, 2)
        row += 1

        form.addWidget(
            self.button(
                "🧾 Xuất phiếu hình ảnh",
                self.export_selected_receipt
            ),
            row,
            0,
            1,
            2
        )

        row += 1

        form.addWidget(
            self.button(
                "↻ Xóa form",
                self.clear_rental_form
            ),
            row,
            0,
            1,
            2
        )

        # RIGHT
        right_group = QGroupBox("XE ĐANG THUÊ")
        right_layout = QVBoxLayout(right_group)

        toolbar = QHBoxLayout()

        toolbar.addWidget(
            self.button("🔄 Làm mới", self.refresh_rentals)
        )

        toolbar.addWidget(
            self.button(
                "⏱ Gia hạn",
                self.extend_selected,
                "warning"
            )
        )

        toolbar.addWidget(
            self.button(
                "↩ Trả xe",
                self.return_selected,
                "success"
            )
        )

        toolbar.addWidget(
            self.button(
                "🧾 Phiếu",
                self.export_selected_receipt
            )
        )

        toolbar.addStretch()

        right_layout.addLayout(toolbar)

        columns = [
            "Đơn",
            "Mã xe",
            "Biển số",
            "Khách hàng",
            "Điện thoại",
            "Ngày giờ giao",
            "Hạn trả",
            "Giá/24h",
            "Cọc",
            "Trạng thái",
        ]

        self.rental_table = QTableWidget(0, len(columns))
        self.rental_table.setHorizontalHeaderLabels(columns)
        self.prepare_table(self.rental_table)
        self.rental_table.itemSelectionChanged.connect(
            self.rental_selected
        )

        right_layout.addWidget(self.rental_table)

        root.addWidget(left_group)
        root.addWidget(right_group, 1)

        self.tabs.addTab(tab, "GIAO / TRẢ XE")

    # ========================================================
    # VEHICLE TAB
    # ========================================================

    def build_vehicle_tab(self):

        tab = QWidget()
        layout = QVBoxLayout(tab)

        form_group = QGroupBox("THÊM XE")
        form = QGridLayout(form_group)

        self.v_code = QLineEdit()
        self.v_plate = QLineEdit()
        self.v_model = QLineEdit()
        self.v_color = QLineEdit()
        self.v_price = QLineEdit()

        fields = [
            ("Mã xe", self.v_code),
            ("Biển số", self.v_plate),
            ("Dòng xe", self.v_model),
            ("Màu", self.v_color),
            ("Giá / 24h", self.v_price),
        ]

        for i, (text, widget) in enumerate(fields):

            form.addWidget(
                self.label(text),
                0,
                i * 2
            )

            form.addWidget(
                widget,
                0,
                i * 2 + 1
            )

        form.addWidget(
            self.button(
                "➕ Thêm xe",
                self.add_vehicle,
                "success"
            ),
            1,
            0,
            1,
            2
        )

        form.addWidget(
            self.button(
                "🗑 Xóa xe",
                self.delete_vehicle,
                "danger"
            ),
            1,
            2,
            1,
            2
        )

        layout.addWidget(form_group)

        columns = [
            "Mã xe",
            "Biển số",
            "Dòng xe",
            "Màu",
            "Giá / 24h",
            "Trạng thái",
            "Ghi chú",
        ]

        self.vehicle_table = QTableWidget(0, len(columns))
        self.vehicle_table.setHorizontalHeaderLabels(columns)
        self.prepare_table(self.vehicle_table)

        layout.addWidget(self.vehicle_table)

        self.tabs.addTab(tab, "XE")

    # ========================================================
    # EXPENSE TAB
    # ========================================================

    def build_expense_tab(self):

        tab = QWidget()
        layout = QVBoxLayout(tab)

        group = QGroupBox("NHẬP CHI PHÍ")
        form = QGridLayout(group)

        self.e_date = QLineEdit()
        self.e_date.setText(now().strftime("%d/%m/%Y"))

        self.e_category = QLineEdit()

        self.e_amount = QLineEdit()
        self.e_amount.setText("0")

        self.e_note = QLineEdit()

        form.addWidget(self.label("Ngày"), 0, 0)
        form.addWidget(self.e_date, 0, 1)

        form.addWidget(self.label("Loại"), 0, 2)
        form.addWidget(self.e_category, 0, 3)

        form.addWidget(self.label("Số tiền"), 0, 4)
        form.addWidget(self.e_amount, 0, 5)

        form.addWidget(self.label("Ghi chú"), 1, 0)
        form.addWidget(self.e_note, 1, 1, 1, 5)

        form.addWidget(
            self.button(
                "➕ Lưu chi phí",
                self.add_expense,
                "success"
            ),
            1,
            6
        )

        layout.addWidget(group)

        columns = [
            "Ngày",
            "Loại",
            "Số tiền",
            "Ghi chú",
        ]

        self.expense_table = QTableWidget(0, len(columns))
        self.expense_table.setHorizontalHeaderLabels(columns)
        self.prepare_table(self.expense_table)

        layout.addWidget(self.expense_table)

        self.tabs.addTab(tab, "CHI PHÍ")

    # ========================================================
    # REPORT TAB
    # ========================================================

    def build_report_tab(self):

        tab = QWidget()
        layout = QVBoxLayout(tab)

        toolbar = QHBoxLayout()

        toolbar.addWidget(self.label("Tháng"))

        self.report_month = QComboBox()

        for i in range(1, 13):
            self.report_month.addItem(f"{i:02d}", i)

        self.report_month.setCurrentIndex(now().month - 1)

        toolbar.addWidget(self.report_month)

        toolbar.addWidget(self.label("Năm"))

        self.report_year = QLineEdit()
        self.report_year.setText(str(now().year))
        self.report_year.setFixedWidth(90)

        toolbar.addWidget(self.report_year)

        toolbar.addWidget(
            self.button("📊 Xem", self.refresh_report)
        )

        toolbar.addWidget(
            self.button(
                "📗 Xuất Excel",
                self.export_excel,
                "success"
            )
        )

        toolbar.addStretch()

        layout.addLayout(toolbar)

        self.report_text = QTextEdit()
        self.report_text.setReadOnly(True)

        layout.addWidget(self.report_text)

        self.tabs.addTab(tab, "DOANH THU / EXCEL")

    # ========================================================
    # VEHICLES
    # ========================================================

    def refresh_vehicles(self):

        rows = self.db.fetchall(
            "SELECT * FROM vehicles ORDER BY code"
        )

        self.vehicle_table.setRowCount(0)

        for r in rows:

            row = self.vehicle_table.rowCount()
            self.vehicle_table.insertRow(row)

            values = [
                r["code"],
                r["plate"],
                r["model"],
                r["color"] or "",
                money(r["default_price"]),
                r["status"],
                r["note"] or "",
            ]

            for col, value in enumerate(values):

                color = "#111827"

                if r["status"] != "Trống":
                    color = "#92400e"

                self.set_item(
                    self.vehicle_table,
                    row,
                    col,
                    value,
                    color
                )

    def refresh_vehicle_combo(self):

        self.vehicle_combo.blockSignals(True)
        self.vehicle_combo.clear()

        rows = self.db.fetchall("""
            SELECT *
            FROM vehicles
            WHERE status='Trống'
            ORDER BY code
        """)

        self.vehicle_combo.addItem("-- Chọn xe --", None)

        for r in rows:

            text = (
                f"{r['code']} | "
                f"{r['plate']} | "
                f"{r['model']}"
            )

            self.vehicle_combo.addItem(
                text,
                r["id"]
            )

        self.vehicle_combo.blockSignals(False)

    def vehicle_selected(self, index):

        vehicle_id = self.vehicle_combo.currentData()

        if not vehicle_id:
            return

        r = self.db.fetchone(
            "SELECT * FROM vehicles WHERE id=?",
            (vehicle_id,)
        )

        if r:
            self.r_price.setText(
                str(int(num(r["default_price"])))
            )

    def add_vehicle(self):

        try:

            code = self.v_code.text().strip()
            plate = self.v_plate.text().strip()
            model = self.v_model.text().strip()
            color = self.v_color.text().strip()

            price = parse_money(
                self.v_price.text()
            )

            if not code:
                raise ValueError("Chưa nhập mã xe.")

            if not plate:
                raise ValueError("Chưa nhập biển số.")

            if not model:
                raise ValueError("Chưa nhập dòng xe.")

            self.db.execute("""
                INSERT INTO vehicles
                (code,plate,model,color,default_price,status)
                VALUES(?,?,?,?,?,'Trống')
            """, (
                code,
                plate,
                model,
                color,
                price
            ))

            for widget in [
                self.v_code,
                self.v_plate,
                self.v_model,
                self.v_color,
                self.v_price,
            ]:
                widget.clear()

            self.refresh_all()

            self.show_info(
                "Thành công",
                "Đã thêm xe."
            )

        except sqlite3.IntegrityError:
            self.show_error(
                "Mã xe hoặc biển số đã tồn tại."
            )

        except Exception as e:
            self.show_error(e)

    def delete_vehicle(self):

        row = self.vehicle_table.currentRow()

        if row < 0:
            return self.show_warning(
                "Hãy chọn xe cần xóa."
            )

        code = self.vehicle_table.item(row, 0).text()

        vehicle = self.db.fetchone(
            "SELECT * FROM vehicles WHERE code=?",
            (code,)
        )

        if not vehicle:
            return

        if vehicle["status"] != "Trống":

            return self.show_warning(
                "Xe đang được thuê, không thể xóa."
            )

        answer = QMessageBox.question(
            self,
            "Xác nhận",
            f"Bạn có chắc muốn xóa xe {code}?"
        )

        if answer == QMessageBox.StandardButton.Yes:

            self.db.execute(
                "DELETE FROM vehicles WHERE id=?",
                (vehicle["id"],)
            )

            self.refresh_all()

    # ========================================================
    # CREATE RENTAL
    # ========================================================

    def get_start_datetime(self):

        d = self.r_date.date()
        t = self.r_time.time()

        return datetime(
            d.year(),
            d.month(),
            d.day(),
            t.hour(),
            t.minute(),
        )

    def create_rental(self):

        try:

            name = self.r_name.text().strip()
            phone = self.r_phone.text().strip()
            id_number = self.r_id.text().strip()
            address = self.r_address.text().strip()

            if not name:
                raise ValueError(
                    "Chưa nhập tên khách hàng."
                )

            if not phone:
                raise ValueError(
                    "Chưa nhập số điện thoại."
                )

            vehicle_id = self.vehicle_combo.currentData()

            if not vehicle_id:
                raise ValueError(
                    "Chưa chọn xe."
                )

            vehicle = self.db.fetchone(
                "SELECT * FROM vehicles WHERE id=?",
                (vehicle_id,)
            )

            if not vehicle:
                raise ValueError(
                    "Không tìm thấy xe."
                )

            if vehicle["status"] != "Trống":
                raise ValueError(
                    "Xe này hiện không còn trống."
                )

            price = parse_money(
                self.r_price.text()
            )

            days = self.r_days.value()

            deposit = parse_money(
                self.r_deposit.text()
            )

            start = self.get_start_datetime()

            due = start + timedelta(
                hours=24 * days
            )

            note = self.r_note.toPlainText().strip()

            # CUSTOMER
            customer = self.db.fetchone("""
                SELECT *
                FROM customers
                WHERE phone=?
                ORDER BY id DESC
                LIMIT 1
            """, (phone,))

            if customer:

                customer_id = customer["id"]

                self.db.execute("""
                    UPDATE customers
                    SET name=?,
                        id_number=?,
                        address=?,
                        note=?
                    WHERE id=?
                """, (
                    name,
                    id_number,
                    address,
                    note,
                    customer_id
                ))

            else:

                cur = self.db.execute("""
                    INSERT INTO customers
                    (name,phone,id_number,address,note,created_at)
                    VALUES(?,?,?,?,?,?)
                """, (
                    name,
                    phone,
                    id_number,
                    address,
                    note,
                    dt_to_str(now())
                ))

                customer_id = cur.lastrowid

            total = price * days

            cur = self.db.execute("""
                INSERT INTO rentals(
                    customer_id,
                    vehicle_id,
                    booking_time,
                    start_time,
                    planned_days,
                    due_time,
                    price_per_day,
                    deposit,
                    rental_total,
                    final_total,
                    status,
                    note,
                    delivery_note
                )
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                customer_id,
                vehicle_id,
                dt_to_str(now()),
                dt_to_str(start),
                days,
                dt_to_str(due),
                price,
                deposit,
                total,
                total,
                "Đang thuê",
                note,
                note,
            ))

            rental_id = cur.lastrowid

            self.db.execute("""
                UPDATE vehicles
                SET status='Đang thuê'
                WHERE id=?
            """, (vehicle_id,))

            self.selected_rental = rental_id

            receipt = None

            if PIL_OK:
                try:
                    receipt = self.create_receipt_image(
                        rental_id
                    )
                except Exception:
                    receipt = None

            self.refresh_all()

            message = (
                f"Đơn #{rental_id}\n\n"
                f"Xe: {vehicle['code']} - "
                f"{vehicle['plate']}\n"
                f"Khách: {name}\n\n"
                f"Giao xe:\n{dt_to_str(start)}\n\n"
                f"Hạn trả:\n{dt_to_str(due)}\n\n"
                f"Tiền thuê:\n{money(total)}\n\n"
                f"Tiền cọc:\n{money(deposit)}"
            )

            if receipt:
                message += f"\n\nPhiếu:\n{receipt}"

            self.show_info(
                "GIAO XE THÀNH CÔNG",
                message
            )

            self.clear_rental_form()

        except Exception as e:
            self.show_error(
                f"Không thể giao xe:\n{e}"
            )

    # ========================================================
    # RENTALS
    # ========================================================

    def active_rentals(self):

        return self.db.fetchall("""
            SELECT
                r.*,

                c.name AS customer_name,
                c.phone AS customer_phone,
                c.id_number AS customer_id_number,
                c.address AS customer_address,

                v.code AS vehicle_code,
                v.plate AS vehicle_plate,
                v.model AS vehicle_model,
                v.color AS vehicle_color

            FROM rentals r

            LEFT JOIN customers c
                ON c.id=r.customer_id

            LEFT JOIN vehicles v
                ON v.id=r.vehicle_id

            WHERE r.status IN ('Đang thuê','Gia hạn')

            ORDER BY r.due_time
        """)

    def is_overdue(self, rental):

        try:

            due = parse_datetime(
                rental["due_time"]
            )

            return (
                due < now()
                and rental["status"]
                in ("Đang thuê", "Gia hạn")
            )

        except Exception:
            return False

    def refresh_rentals(self):

        rows = self.active_rentals()

        self.rental_table.setRowCount(0)

        for r in rows:

            row = self.rental_table.rowCount()
            self.rental_table.insertRow(row)

            overdue = self.is_overdue(r)

            status = (
                "⚠ QUÁ HẠN"
                if overdue
                else r["status"]
            )

            values = [
                r["id"],
                r["vehicle_code"],
                r["vehicle_plate"],
                r["customer_name"] or "",
                r["customer_phone"] or "",
                r["start_time"],
                r["due_time"],
                money(r["price_per_day"]),
                money(r["deposit"]),
                status,
            ]

            for col, value in enumerate(values):

                color = "#991b1b" if overdue else "#111827"

                self.set_item(
                    self.rental_table,
                    row,
                    col,
                    value,
                    color
                )

                if overdue:

                    self.rental_table.item(
                        row,
                        col
                    ).setBackground(
                        QColor("#fecaca")
                    )

                elif col == 9:

                    self.rental_table.item(
                        row,
                        col
                    ).setBackground(
                        QColor("#dcfce7")
                    )

        self.refresh_vehicle_combo()

    def rental_selected(self):

        row = self.rental_table.currentRow()

        if row < 0:
            self.selected_rental = None
            return

        try:
            self.selected_rental = int(
                self.rental_table.item(
                    row,
                    0
                ).text()
            )
        except Exception:
            self.selected_rental = None

    # ========================================================
    # EXTEND
    # ========================================================

    def extend_selected(self):

        if not self.selected_rental:

            return self.show_warning(
                "Hãy chọn đơn cần gia hạn."
            )

        r = self.db.fetchone(
            "SELECT * FROM rentals WHERE id=?",
            (self.selected_rental,)
        )

        if not r:
            return

        if r["status"] not in (
            "Đang thuê",
            "Gia hạn"
        ):

            return self.show_warning(
                "Đơn này đã trả xe."
            )

        dialog = QDialog(self)
        dialog.setWindowTitle(
            f"GIA HẠN ĐƠN #{r['id']}"
        )
        dialog.resize(400, 220)

        layout = QFormLayout(dialog)

        due_label = QLabel(
            f"Hạn hiện tại: {r['due_time']}"
        )
        due_label.setStyleSheet(
            "color:#1d4ed8;font-weight:bold;"
        )

        layout.addRow(due_label)

        days = QSpinBox()
        days.setRange(1, 365)
        days.setValue(1)

        price = QLineEdit()
        price.setText(
            str(int(num(r["price_per_day"])))
        )

        layout.addRow(
            "Số ngày gia hạn",
            days
        )

        layout.addRow(
            "Giá / 24 giờ",
            price
        )

        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Save
            | QDialogButtonBox.StandardButton.Cancel
        )

        buttons.accepted.connect(
            dialog.accept
        )
        buttons.rejected.connect(
            dialog.reject
        )

        layout.addRow(buttons)

        if dialog.exec() != QDialog.DialogCode.Accepted:
            return

        try:

            d = days.value()
            p = parse_money(price.text())

            old = parse_datetime(
                r["due_time"]
            )

            new = old + timedelta(
                hours=24 * d
            )

            added = p * d

            self.db.execute("""
                INSERT INTO extensions(
                    rental_id,
                    old_due,
                    new_due,
                    added_days,
                    added_amount,
                    created_at
                )
                VALUES(?,?,?,?,?,?)
            """, (
                r["id"],
                r["due_time"],
                dt_to_str(new),
                d,
                added,
                dt_to_str(now()),
            ))

            self.db.execute("""
                UPDATE rentals
                SET planned_days=planned_days+?,
                    due_time=?,
                    rental_total=rental_total+?,
                    final_total=final_total+?,
                    status='Gia hạn'
                WHERE id=?
            """, (
                d,
                dt_to_str(new),
                added,
                added,
                r["id"],
            ))

            self.refresh_all()

            self.show_info(
                "ĐÃ GIA HẠN",
                f"Hạn mới:\n{dt_to_str(new)}\n\n"
                f"Tiền gia hạn:\n{money(added)}"
            )

        except Exception as e:
            self.show_error(e)

    # ========================================================
    # RETURN
    # ========================================================

    def return_selected(self):

        if not self.selected_rental:

            return self.show_warning(
                "Hãy chọn đơn cần trả."
            )

        r = self.db.fetchone("""
            SELECT
                r.*,
                c.name AS customer_name,
                c.phone AS customer_phone,
                v.code,
                v.plate,
                v.model
            FROM rentals r
            LEFT JOIN customers c
                ON c.id=r.customer_id
            LEFT JOIN vehicles v
                ON v.id=r.vehicle_id
            WHERE r.id=?
        """, (self.selected_rental,))

        if not r:
            return

        dialog = QDialog(self)
        dialog.setWindowTitle(
            f"TRẢ XE - ĐƠN #{r['id']}"
        )
        dialog.resize(450, 350)

        layout = QFormLayout(dialog)

        return_time = QLineEdit(
            dt_to_str(now())
        )

        refund = QLineEdit("0")
        damage = QLineEdit("0")
        fuel = QLineEdit("0")
        discount = QLineEdit("0")

        layout.addRow(
            "Ngày giờ trả xe",
            return_time
        )

        layout.addRow(
            "Hoàn tiền khách",
            refund
        )

        layout.addRow(
            "Phí hư hỏng",
            damage
        )

        layout.addRow(
            "Phí xăng",
            fuel
        )

        layout.addRow(
            "Giảm giá",
            discount
        )

        note = QLabel(
            "Hoàn tiền khách sẽ được trừ khỏi tổng tiền."
        )

        note.setStyleSheet(
            "color:#b45309;font-weight:bold;"
        )

        layout.addRow(note)

        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok
            | QDialogButtonBox.StandardButton.Cancel
        )

        buttons.accepted.connect(
            dialog.accept
        )
        buttons.rejected.connect(
            dialog.reject
        )

        layout.addRow(buttons)

        if dialog.exec() != QDialog.DialogCode.Accepted:
            return

        try:

            return_dt = parse_datetime(
                return_time.text()
            )

            start_dt = parse_datetime(
                r["start_time"]
            )

            if return_dt < start_dt:
                raise ValueError(
                    "Ngày giờ trả không được trước ngày giao."
                )

            refund_value = parse_money(
                refund.text()
            )

            damage_value = parse_money(
                damage.text()
            )

            fuel_value = parse_money(
                fuel.text()
            )

            discount_value = parse_money(
                discount.text()
            )

            total = (
                num(r["rental_total"])
                + damage_value
                + fuel_value
                - discount_value
                - refund_value
            )

            if total < 0:
                raise ValueError(
                    "Tổng tiền không thể nhỏ hơn 0."
                )

            self.db.execute("""
                UPDATE rentals
                SET actual_return=?,
                    refund=?,
                    damage_fee=?,
                    fuel_fee=?,
                    discount=?,
                    final_total=?,
                    status='Đã trả'
                WHERE id=?
            """, (
                dt_to_str(return_dt),
                refund_value,
                damage_value,
                fuel_value,
                discount_value,
                total,
                r["id"],
            ))

            self.db.execute("""
                UPDATE vehicles
                SET status='Trống'
                WHERE id=?
            """, (r["vehicle_id"],))

            self.refresh_all()

            self.show_info(
                "TRẢ XE THÀNH CÔNG",
                f"Đơn #{r['id']}\n\n"
                f"Xe: {r['code']} - {r['plate']}\n"
                f"Ngày giờ trả: {dt_to_str(return_dt)}\n\n"
                f"Hoàn tiền: {money(refund_value)}\n"
                f"Phí hư hỏng: {money(damage_value)}\n"
                f"Phí xăng: {money(fuel_value)}\n"
                f"Giảm giá: {money(discount_value)}\n\n"
                f"TỔNG TIỀN: {money(total)}"
            )

            self.selected_rental = None

        except Exception as e:
            self.show_error(
                f"Lỗi trả xe:\n{e}"
            )

    # ========================================================
    # RENTAL DETAIL
    # ========================================================

    def rental_full(self, rental_id):

        return self.db.fetchone("""
            SELECT
                r.*,

                c.name AS customer_name,
                c.phone AS customer_phone,
                c.id_number AS customer_id_number,
                c.address AS customer_address,

                v.code AS vehicle_code,
                v.plate AS vehicle_plate,
                v.model AS vehicle_model,
                v.color AS vehicle_color

            FROM rentals r

            LEFT JOIN customers c
                ON c.id=r.customer_id

            LEFT JOIN vehicles v
                ON v.id=r.vehicle_id

            WHERE r.id=?
        """, (rental_id,))

    # ========================================================
    # RECEIPT PNG
    # ========================================================

    def find_font(self, bold=False):

        if bold:
            candidates = [
                "C:/Windows/Fonts/arialbd.ttf",
                "C:/Windows/Fonts/segoeuib.ttf",
                "C:/Windows/Fonts/tahomabd.ttf",
            ]
        else:
            candidates = [
                "C:/Windows/Fonts/arial.ttf",
                "C:/Windows/Fonts/segoeui.ttf",
                "C:/Windows/Fonts/tahoma.ttf",
            ]

        for path in candidates:

            if os.path.exists(path):
                return path

        return None

    def create_receipt_image(self, rental_id):

        if not PIL_OK:
            raise ValueError(
                "Chưa cài Pillow.\n\n"
                "pip install Pillow"
            )

        r = self.rental_full(rental_id)

        if not r:
            raise ValueError(
                "Không tìm thấy đơn."
            )

        start_dt = parse_datetime(
            r["start_time"]
        )

        date_str = start_dt.strftime(
            "%d-%m-%Y"
        )

        filename = (
            f"PHIEU GIAO XE SO "
            f"{rental_id}_{date_str}.png"
        )

        path = os.path.join(
            PHIEU_DIR,
            filename
        )

        width = 1100
        height = 1850

        image = Image.new(
            "RGB",
            (width, height),
            "white"
        )

        draw = ImageDraw.Draw(image)

        normal_font = self.find_font(False)
        bold_font = self.find_font(True)

        if normal_font:

            title_font = ImageFont.truetype(
                bold_font or normal_font,
                34
            )

            head_font = ImageFont.truetype(
                bold_font or normal_font,
                23
            )

            font = ImageFont.truetype(
                normal_font,
                20
            )

            small = ImageFont.truetype(
                normal_font,
                17
            )

        else:

            title_font = ImageFont.load_default()
            head_font = ImageFont.load_default()
            font = ImageFont.load_default()
            small = ImageFont.load_default()

        x = 60
        y = 35

        draw.text(
            (x, y),
            "PHIẾU GIAO XE MÁY",
            font=title_font,
            fill="#172033"
        )

        y += 60

        draw.text(
            (x, y),
            f"PHIẾU SỐ {rental_id} - {date_str}",
            font=head_font,
            fill="#334155"
        )

        y += 45

        draw.line(
            (x, y, width - x, y),
            fill="#94a3b8",
            width=2
        )

        y += 30

        sections = [
            (
                "KHÁCH HÀNG",
                [
                    ("Tên", r["customer_name"] or ""),
                    ("Điện thoại", r["customer_phone"] or ""),
                    (
                        "CCCD/Hộ chiếu",
                        r["customer_id_number"] or ""
                    ),
                    (
                        "Địa chỉ",
                        r["customer_address"] or ""
                    ),
                ]
            ),
            (
                "XE",
                [
                    ("Mã xe", r["vehicle_code"]),
                    ("Biển số", r["vehicle_plate"]),
                    ("Dòng xe", r["vehicle_model"]),
                    ("Màu", r["vehicle_color"] or ""),
                ]
            ),
            (
                "THỜI GIAN",
                [
                    ("Ngày giờ giao", r["start_time"]),
                    ("Ngày giờ hết hạn", r["due_time"]),
                    (
                        "Số ngày",
                        f"{r['planned_days']} x 24 giờ"
                    ),
                ]
            ),
        ]

        for section_title, lines in sections:

            draw.text(
                (x, y),
                section_title,
                font=head_font,
                fill="#1d4ed8"
            )

            y += 38

            for label, value in lines:

                draw.text(
                    (x + 15, y),
                    label,
                    font=small,
                    fill="#64748b"
                )

                draw.text(
                    (x + 270, y),
                    str(value),
                    font=font,
                    fill="#111827"
                )

                y += 38

            y += 10

        draw.line(
            (x, y, width - x, y),
            fill="#cbd5e1",
            width=2
        )

        y += 25

        draw.text(
            (x, y),
            "THANH TOÁN",
            font=head_font,
            fill="#15803d"
        )

        y += 42

        payment = [
            (
                "Giá / 24 giờ",
                money(r["price_per_day"])
            ),
            (
                "Tiền cọc",
                money(r["deposit"])
            ),
            (
                "TỔNG TIỀN",
                money(r["rental_total"])
            ),
        ]

        for label, value in payment:

            draw.text(
                (x + 15, y),
                label,
                font=font,
                fill="#111827"
            )

            draw.text(
                (x + 650, y),
                value,
                font=font,
                fill="#111827"
            )

            y += 40

        y += 15

        draw.line(
            (x, y, width - x, y),
            fill="#cbd5e1",
            width=2
        )

        y += 25

        draw.text(
            (x, y),
            "GHI CHÚ",
            font=head_font,
            fill="#334155"
        )

        y += 38

        note = r["delivery_note"] or ""

        if note:

            for i in range(0, len(note), 75):

                draw.text(
                    (x + 15, y),
                    note[i:i + 75],
                    font=font,
                    fill="#111827"
                )

                y += 30

        else:

            draw.text(
                (x + 15, y),
                "Không có",
                font=font,
                fill="#64748b"
            )

            y += 30

        y += 20

        draw.text(
            (x, y),
            "LƯU Ý",
            font=head_font,
            fill="#b91c1c"
        )

        y += 38

        notes = [
            "• Thời gian thuê tính theo 24 giờ kể từ lúc giao xe.",
            "• Đơn quá hạn sẽ tự động hiển thị màu đỏ.",
        ]

        for text in notes:

            draw.text(
                (x + 15, y),
                text,
                font=small,
                fill="#334155"
            )

            y += 30

        y += 35

        draw.text(
            (x, y),
            "Nhân viên: __________________________",
            font=font,
            fill="#111827"
        )

        draw.text(
            (x + 570, y),
            "Khách hàng: __________________",
            font=font,
            fill="#111827"
        )

        y += 60

        draw.line(
            (x, y, width - x, y),
            fill="#94a3b8",
            width=2
        )

        y += 25

        # QR
        if os.path.exists(QR_FILE):

            try:

                qr = Image.open(
                    QR_FILE
                ).convert("RGB")

                qr.thumbnail(
                    (280, 280),
                    Image.Resampling.LANCZOS
                )

                text = "MÃ QR THANH TOÁN"

                bbox = draw.textbbox(
                    (0, 0),
                    text,
                    font=head_font
                )

                draw.text(
                    (
                        (width - (bbox[2] - bbox[0])) // 2,
                        y
                    ),
                    text,
                    font=head_font,
                    fill="#1d4ed8"
                )

                qr_x = (
                    width - qr.width
                ) // 2

                qr_y = y + 45

                image.paste(
                    qr,
                    (qr_x, qr_y)
                )

                text2 = "Quét mã QR để thanh toán"

                bbox2 = draw.textbbox(
                    (0, 0),
                    text2,
                    font=small
                )

                draw.text(
                    (
                        (width - (bbox2[2] - bbox2[0])) // 2,
                        qr_y + qr.height + 15
                    ),
                    text2,
                    font=small,
                    fill="#475569"
                )

            except Exception:

                draw.text(
                    (x, y),
                    "QR_THANH_TOAN.png không đọc được.",
                    font=font,
                    fill="#b91c1c"
                )

        else:

            text = "CHƯA CÓ QR THANH TOÁN"

            bbox = draw.textbbox(
                (0, 0),
                text,
                font=head_font
            )

            draw.text(
                (
                    (width - (bbox[2] - bbox[0])) // 2,
                    y + 20
                ),
                text,
                font=head_font,
                fill="#b91c1c"
            )

            text2 = (
                "Đặt file QR_THANH_TOAN.png "
                "cạnh file chương trình."
            )

            bbox2 = draw.textbbox(
                (0, 0),
                text2,
                font=small
            )

            draw.text(
                (
                    (width - (bbox2[2] - bbox2[0])) // 2,
                    y + 65
                ),
                text2,
                font=small,
                fill="#64748b"
            )

        image.save(path)

        return path

    def export_selected_receipt(self):

        if not self.selected_rental:

            return self.show_warning(
                "Hãy chọn một đơn."
            )

        try:

            path = self.create_receipt_image(
                self.selected_rental
            )

            self.show_info(
                "Đã tạo phiếu",
                f"Phiếu giao xe đã được tạo:\n\n{path}"
            )

        except Exception as e:
            self.show_error(e)

    # ========================================================
    # CLEAR RENTAL FORM
    # ========================================================

    def clear_rental_form(self):

        for widget in [
            self.r_name,
            self.r_phone,
            self.r_id,
            self.r_address,
            self.r_price,
        ]:
            widget.clear()

        self.vehicle_combo.setCurrentIndex(0)

        self.r_days.setValue(1)

        self.r_deposit.setText("0")

        self.r_date.setDate(
            QDate.currentDate()
        )

        self.r_time.setTime(
            QTime.currentTime()
        )

        self.r_note.clear()

    # ========================================================
    # EXPENSES
    # ========================================================

    def add_expense(self):

        try:

            date = self.e_date.text().strip()

            datetime.strptime(
                date,
                "%d/%m/%Y"
            )

            category = (
                self.e_category.text().strip()
                or "Khác"
            )

            amount = parse_money(
                self.e_amount.text()
            )

            note = self.e_note.text().strip()

            self.db.execute("""
                INSERT INTO expenses(
                    expense_date,
                    category,
                    amount,
                    note
                )
                VALUES(?,?,?,?)
            """, (
                date,
                category,
                amount,
                note,
            ))

            self.e_amount.setText("0")
            self.e_note.clear()

            self.refresh_expenses()
            self.refresh_report()

        except Exception as e:
            self.show_error(e)

    def refresh_expenses(self):

        rows = self.db.fetchall("""
            SELECT *
            FROM expenses
            ORDER BY id DESC
        """)

        self.expense_table.setRowCount(0)

        for r in rows:

            row = self.expense_table.rowCount()
            self.expense_table.insertRow(row)

            values = [
                r["expense_date"],
                r["category"],
                money(r["amount"]),
                r["note"] or "",
            ]

            for col, value in enumerate(values):

                self.set_item(
                    self.expense_table,
                    row,
                    col,
                    value
                )

    # ========================================================
    # REPORT DATA
    # ========================================================

    def report_data(self, month, year):

        rows = self.db.fetchall("""
            SELECT
                r.*,
                c.name AS customer_name,
                c.phone,
                v.code,
                v.plate,
                v.model

            FROM rentals r

            LEFT JOIN customers c
                ON c.id=r.customer_id

            LEFT JOIN vehicles v
                ON v.id=r.vehicle_id

            WHERE r.status='Đã trả'

            AND substr(
                r.actual_return,
                7,
                4
            )=?

            AND substr(
                r.actual_return,
                4,
                2
            )=?

            ORDER BY r.actual_return
        """, (
            str(year),
            f"{month:02d}"
        ))

        expenses = self.db.fetchall("""
            SELECT *
            FROM expenses

            WHERE substr(
                expense_date,
                7,
                4
            )=?

            AND substr(
                expense_date,
                4,
                2
            )=?

            ORDER BY expense_date
        """, (
            str(year),
            f"{month:02d}"
        ))

        return rows, expenses

    # ========================================================
    # REPORT
    # ========================================================

    def refresh_report(self):

        try:

            month = self.report_month.currentData()

            if not month:
                month = self.report_month.currentIndex() + 1

            year = int(
                self.report_year.text()
            )

        except Exception:
            return

        rows, expenses = self.report_data(
            month,
            year
        )

        revenue = sum(
            num(r["final_total"])
            for r in rows
        )

        cost = sum(
            num(e["amount"])
            for e in expenses
        )

        profit = revenue - cost

        text = []

        text.append(
            f"DOANH THU THÁNG "
            f"{month:02d}/{year}"
        )

        text.append("=" * 90)
        text.append("")

        text.append(
            f"Số đơn đã trả : {len(rows)}"
        )

        text.append(
            f"DOANH THU     : {money(revenue)}"
        )

        text.append(
            f"CHI PHÍ       : {money(cost)}"
        )

        text.append(
            f"LỢI NHUẬN     : {money(profit)}"
        )

        text.append("")
        text.append("CHI TIẾT ĐƠN")
        text.append("-" * 90)

        for r in rows:

            text.append(
                f"#{r['id']} | "
                f"{r['actual_return']} | "
                f"{r['code']} | "
                f"{r['plate']} | "
                f"{r['customer_name'] or ''} | "
                f"{money(r['final_total'])}"
            )

        text.append("")
        text.append("CHI PHÍ")
        text.append("-" * 90)

        for e in expenses:

            text.append(
                f"{e['expense_date']} | "
                f"{e['category']} | "
                f"{money(e['amount'])} | "
                f"{e['note'] or ''}"
            )

        self.report_text.setPlainText(
            "\n".join(text)
        )

    # ========================================================
    # EXCEL
    # ========================================================

    def export_excel(self):

        if not XLSX_OK:

            return self.show_error(
                "Chưa cài openpyxl.\n\n"
                "pip install openpyxl"
            )

        try:

            month = self.report_month.currentData()

            if not month:
                month = self.report_month.currentIndex() + 1

            year = int(
                self.report_year.text()
            )

            rows, expenses = self.report_data(
                month,
                year
            )

            wb = Workbook()

            ws = wb.active
            ws.title = "Doanh thu"

            headers = [
                "Đơn",
                "Ngày trả",
                "Mã xe",
                "Biển số",
                "Dòng xe",
                "Khách hàng",
                "Điện thoại",
                "Ngày giờ giao",
                "Hạn trả",
                "Giá / 24h (VNĐ)",
                "Cọc (VNĐ)",
                "Hoàn tiền khách (VNĐ)",
                "Phí hư hỏng (VNĐ)",
                "Xăng (VNĐ)",
                "Giảm giá (VNĐ)",
                "Tổng tiền (VNĐ)",
                "Trạng thái",
            ]

            ws.append(headers)

            for cell in ws[1]:

                cell.font = XLFont(
                    bold=True,
                    color="FFFFFF"
                )

                cell.fill = PatternFill(
                    "solid",
                    fgColor="1F4E78"
                )

                cell.alignment = Alignment(
                    horizontal="center"
                )

            for r in rows:

                ws.append([
                    r["id"],
                    r["actual_return"],
                    r["code"],
                    r["plate"],
                    r["model"],
                    r["customer_name"],
                    r["phone"],
                    r["start_time"],
                    r["due_time"],
                    r["price_per_day"],
                    r["deposit"],
                    r["refund"],
                    r["damage_fee"],
                    r["fuel_fee"],
                    r["discount"],
                    r["final_total"],
                    r["status"],
                ])

            for row in ws.iter_rows(
                min_row=2
            ):

                for cell in row:

                    if 10 <= cell.column <= 16:

                        cell.number_format = (
                            '#,##0" đ"'
                        )

            for col in ws.columns:

                length = max(
                    len(str(c.value or ""))
                    for c in col
                )

                ws.column_dimensions[
                    get_column_letter(
                        col[0].column
                    )
                ].width = min(
                    max(length + 2, 12),
                    30
                )

            ws.freeze_panes = "A2"

            # CHI PHÍ
            ex = wb.create_sheet("Chi phí")

            ex.append([
                "Ngày",
                "Loại",
                "Số tiền (VNĐ)",
                "Ghi chú",
            ])

            for cell in ex[1]:

                cell.font = XLFont(
                    bold=True,
                    color="FFFFFF"
                )

                cell.fill = PatternFill(
                    "solid",
                    fgColor="7F6000"
                )

            for e in expenses:

                ex.append([
                    e["expense_date"],
                    e["category"],
                    e["amount"],
                    e["note"] or "",
                ])

            for row in ex.iter_rows(
                min_row=2
            ):

                row[2].number_format = (
                    '#,##0" đ"'
                )

            # SUMMARY
            sm = wb.create_sheet(
                "Tổng hợp"
            )

            revenue = sum(
                num(r["final_total"])
                for r in rows
            )

            cost = sum(
                num(e["amount"])
                for e in expenses
            )

            profit = revenue - cost

            sm.append([
                "BÁO CÁO DOANH THU",
                f"{month:02d}/{year}"
            ])

            sm.append([
                "Số đơn đã trả",
                len(rows)
            ])

            sm.append([
                "Doanh thu (VNĐ)",
                revenue
            ])

            sm.append([
                "Chi phí (VNĐ)",
                cost
            ])

            sm.append([
                "Lợi nhuận (VNĐ)",
                profit
            ])

            sm["A1"].font = XLFont(
                bold=True,
                size=15
            )

            for cell in [
                "B3",
                "B4",
                "B5"
            ]:

                sm[cell].number_format = (
                    '#,##0" đ"'
                )

            sm.column_dimensions[
                "A"
            ].width = 25

            sm.column_dimensions[
                "B"
            ].width = 25

            filename = (
                f"Doanh_thu_"
                f"{year}_{month:02d}.xlsx"
            )

            path, _ = QFileDialog.getSaveFileName(
                self,
                "Lưu báo cáo Excel",
                os.path.join(BASE, filename),
                "Excel (*.xlsx)"
            )

            if not path:
                return

            wb.save(path)

            self.show_info(
                "Đã xuất Excel",
                f"Đã lưu:\n\n{path}"
            )

        except Exception as e:
            self.show_error(
                f"Lỗi xuất Excel:\n{e}"
            )

    # ========================================================
    # DASHBOARD REFRESH
    # ========================================================

    def refresh_dashboard(self):

        active = self.active_rentals()

        overdue = [
            r
            for r in active
            if self.is_overdue(r)
        ]

        free = self.db.fetchone("""
            SELECT COUNT(*) AS total
            FROM vehicles
            WHERE status='Trống'
        """)

        free_count = (
            free["total"]
            if free
            else 0
        )

        revenue = self.db.fetchone("""
            SELECT
                COALESCE(
                    SUM(final_total),
                    0
                ) AS total

            FROM rentals

            WHERE status='Đã trả'

            AND substr(
                actual_return,
                7,
                4
            )=?

            AND substr(
                actual_return,
                4,
                2
            )=?
        """, (
            str(now().year),
            f"{now().month:02d}"
        ))

        revenue_value = (
            revenue["total"]
            if revenue
            else 0
        )

        self.card_active.setText(
            str(len(active))
        )

        self.card_overdue.setText(
            str(len(overdue))
        )

        self.card_free.setText(
            str(free_count)
        )

        self.card_revenue.setText(
            money(revenue_value)
        )

        self.dashboard_table.setRowCount(0)

        for r in active:

            row = self.dashboard_table.rowCount()

            self.dashboard_table.insertRow(row)

            late = self.is_overdue(r)

            status = (
                "⚠ QUÁ HẠN"
                if late
                else r["status"]
            )

            values = [
                r["vehicle_code"],
                r["vehicle_plate"],
                r["customer_name"] or "",
                r["customer_phone"] or "",
                r["start_time"],
                r["due_time"],
                money(r["rental_total"]),
                status,
            ]

            for col, value in enumerate(values):

                color = (
                    "#991b1b"
                    if late
                    else "#111827"
                )

                self.set_item(
                    self.dashboard_table,
                    row,
                    col,
                    value,
                    color
                )

                if late:

                    self.dashboard_table.item(
                        row,
                        col
                    ).setBackground(
                        QColor("#fecaca")
                    )

    # ========================================================
    # REFRESH ALL
    # ========================================================

    def refresh_all(self):

        self.refresh_vehicles()
        self.refresh_vehicle_combo()
        self.refresh_rentals()
        self.refresh_expenses()
        self.refresh_dashboard()
        self.refresh_report()

    def auto_refresh(self):

        self.refresh_rentals()
        self.refresh_dashboard()

    # ========================================================
    # CLOSE
    # ========================================================

    def closeEvent(self, event):

        try:
            self.db.close()
        except Exception:
            pass

        event.accept()


# ============================================================
# MAIN
# ============================================================

def main():

    app = QApplication(sys.argv)

    app.setApplicationName(
        "Quản lý cho thuê xe máy"
    )

    app.setStyle("Fusion")

    window = MainWindow()
    window.show()

    sys.exit(
        app.exec()
    )


if __name__ == "__main__":
    main()